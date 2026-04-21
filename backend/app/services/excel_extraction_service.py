"""
Excel Extraction Service — Rent Rolls and T-12 Operating Statements

This service parses Rent Roll and T-12 Excel files using openpyxl and extracts structured data.
Uses a hybrid approach: fuzzy keyword matching (Tier 1) with Claude AI fallback (Tier 2).
Uses direct parsing with intelligent header detection and fuzzy matching via rapidfuzz.
"""
import json
import logging
import re
from datetime import datetime
from typing import Dict, Any, List, Optional, Tuple
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from rapidfuzz import fuzz, process

from app.config import settings
from app.services.t12_taxonomy import T12_TAXONOMY, TAXONOMY_TO_SUMMARY_FIELD

logger = logging.getLogger(__name__)


# ==================== FUZZY MATCHING ENGINE ====================

def match_line_item(raw_label: str, taxonomy: dict, threshold: int = 70) -> Optional[str]:
    """
    Match a raw Excel line item label to a canonical taxonomy key.
    Returns the taxonomy key (e.g. 'gsr', 'noi') or None.

    Uses a 3-tier matching strategy:
    1. Exact abbreviation match
    2. Regex pattern match
    3. Fuzzy string match against all keywords
    """
    cleaned = raw_label.strip().lower()
    cleaned = re.sub(r'^[\d\-\.]+\s*', '', cleaned)  # Strip leading GL codes
    cleaned = re.sub(r'\s+', ' ', cleaned)

    if not cleaned or len(cleaned) < 2:
        return None

    # Skip lines that are clearly section headers or separators
    skip_patterns = [
        r'^[\-=_]+$',  # separator lines
        r'^page\s+\d+',
        r'^total\s*$',  # bare "Total" without context
    ]
    for sp in skip_patterns:
        if re.match(sp, cleaned):
            return None

    # 1. Exact abbreviation match
    for key, entry in taxonomy.items():
        if cleaned in [a.lower() for a in entry["abbreviations"]]:
            return key

    # 2. Regex pattern match
    for key, entry in taxonomy.items():
        for pattern in entry["patterns"]:
            if re.search(pattern, cleaned, re.IGNORECASE):
                return key

    # 3. Fuzzy string match against all keywords
    all_keywords = []
    for key, entry in taxonomy.items():
        for kw in entry["keywords"]:
            all_keywords.append((kw, key))

    keyword_strings = [kw for kw, _ in all_keywords]
    match = process.extractOne(cleaned, keyword_strings, scorer=fuzz.token_sort_ratio)
    if match and match[1] >= threshold:
        matched_keyword = match[0]
        for kw, key in all_keywords:
            if kw == matched_keyword:
                return key

    return None


def detect_t12_layout(worksheet) -> dict:
    """
    Scan the worksheet to detect:
    - Which row has month headers
    - Which column has line item labels
    - Which column has the annual total
    - Whether there are GL code columns
    """
    layout = {
        "header_row": None,
        "label_col": None,
        "total_col": None,
        "month_cols": [],
        "has_gl_codes": False,
        "data_start_row": None,
    }

    month_names = [
        "jan", "feb", "mar", "apr", "may", "jun",
        "jul", "aug", "sep", "oct", "nov", "dec",
        "january", "february", "march", "april", "may", "june",
        "july", "august", "september", "october", "november", "december",
    ]
    total_keywords = [
        "total", "annual", "year", "ytd", "t-12", "t12",
        "trailing", "12 month", "12-month", "twelve month",
        "annualized", "combined",
    ]

    # Scan first 20 rows to find month headers
    for row_idx, row in enumerate(worksheet.iter_rows(max_row=20, values_only=False)):
        month_count = 0
        month_cols_found = []
        for cell in row:
            val = str(cell.value).strip().lower() if cell.value else ""
            # Check for month names or date objects
            if any(m in val for m in month_names):
                month_count += 1
                month_cols_found.append(cell.column - 1)
            elif hasattr(cell.value, 'month'):  # datetime object
                month_count += 1
                month_cols_found.append(cell.column - 1)

        if month_count >= 6:
            layout["header_row"] = row_idx
            layout["month_cols"] = month_cols_found
            layout["data_start_row"] = row_idx + 1

            # Find total column in same row
            for cell in row:
                val = str(cell.value).strip().lower() if cell.value else ""
                if any(tk in val for tk in total_keywords):
                    layout["total_col"] = cell.column - 1
            break

    # Detect label column: find the column with the most text strings
    # in rows after the header
    if layout["data_start_row"] is not None:
        col_text_counts = {}
        for row in worksheet.iter_rows(
            min_row=layout["data_start_row"] + 1,
            max_row=min(layout["data_start_row"] + 40, worksheet.max_row),
            values_only=False
        ):
            for cell in row:
                if cell.value and isinstance(cell.value, str) and len(cell.value.strip()) > 2:
                    col_idx = cell.column - 1
                    if col_idx not in layout["month_cols"] and col_idx != layout.get("total_col"):
                        col_text_counts[col_idx] = col_text_counts.get(col_idx, 0) + 1

        if col_text_counts:
            layout["label_col"] = max(col_text_counts, key=col_text_counts.get)
            if layout["label_col"] > 0:
                layout["has_gl_codes"] = True

    return layout


# ==================== DOCUMENT CLASSIFICATION ====================

def classify_excel_document(filepath: str) -> str:
    """
    Classify an Excel file as 'rent_roll', 't12', or 'unknown'.

    Heuristics:
    - Rent Roll: Look for columns like "Unit", "SQFT", "Lease Start", "Lease End", "Market Rent", "Resident"
    - T-12: Look for monthly column headers (Jan/January, Feb/February, ...) and financial line items
    - Check sheet names too — "Rent Roll" in sheet name is a strong signal

    Returns: "rent_roll", "t12", or "unknown"
    """
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    except Exception as e:
        logger.warning(f"Could not open workbook for classification: {e}")
        return "unknown"

    try:
        # Check sheet names
        sheet_names_lower = [s.lower() for s in wb.sheetnames]

        # Strong signals from sheet names
        for name in sheet_names_lower:
            if any(kw in name for kw in ["rent roll", "rentroll", "rr", "unit"]):
                wb.close()
                return "rent_roll"
            if any(kw in name for kw in ["t12", "t-12", "operating", "financials", "income", "statement"]):
                wb.close()
                return "t12"

        # Check first sheet headers
        ws = wb.worksheets[0] if wb.worksheets else None
        if ws is None:
            wb.close()
            return "unknown"

        # Collect first 15 rows of text
        headers_text = []
        for row in ws.iter_rows(min_row=1, max_row=15, values_only=True):
            for cell in row:
                if cell is not None:
                    headers_text.append(str(cell).lower().strip())

        header_blob = " ".join(headers_text)

        # Rent roll keywords
        rr_keywords = ["unit", "sqft", "sq ft", "lease start", "lease end", "market rent",
                       "resident", "tenant", "move in", "status", "occupied", "vacant", "charge"]
        rr_score = sum(1 for kw in rr_keywords if kw in header_blob)

        # T12 keywords
        t12_keywords = ["january", "february", "march", "april", "may", "june", "july", "august",
                        "september", "october", "november", "december", "jan", "feb", "mar", "apr",
                        "gross potential rent", "vacancy", "noi", "operating expenses", "revenue"]
        t12_score = sum(1 for kw in t12_keywords if kw in header_blob)

        wb.close()

        # Decision
        if rr_score >= 3:
            return "rent_roll"
        elif t12_score >= 3:
            return "t12"
        else:
            return "unknown"

    except Exception as e:
        logger.exception(f"Error classifying document: {e}")
        wb.close()
        return "unknown"


# ==================== DATE PARSING ====================

def parse_date_from_filename(filename: str) -> Optional[datetime]:
    """
    Try to extract a date from the filename.

    Common patterns:
    - "1160_Hammond_RR_1_28_26.xlsx" → 2026-01-28
    - "Property_RentRoll_2026-01-28.xlsx" → 2026-01-28
    - "RR_01282026.xlsx" → 2026-01-28
    - "T12_FY_2025.xlsx" → 2025-12-31 (end of fiscal year)
    - "1160_Hammond_FY_2025_T12.xlsx" → 2025-12-31

    Returns None if no date can be parsed.
    """
    if not filename:
        return None

    # Pattern 1: YYYY-MM-DD or YYYY_MM_DD
    match = re.search(r'(\d{4})[-_](\d{1,2})[-_](\d{1,2})', filename)
    if match:
        try:
            year, month, day = int(match.group(1)), int(match.group(2)), int(match.group(3))
            return datetime(year, month, day)
        except (ValueError, TypeError):
            pass

    # Pattern 2: MM_DD_YY or MM-DD-YY
    match = re.search(r'(\d{1,2})[-_](\d{1,2})[-_](\d{2})', filename)
    if match:
        try:
            month, day, year = int(match.group(1)), int(match.group(2)), int(match.group(3))
            year = 2000 + year if year < 100 else year
            return datetime(year, month, day)
        except (ValueError, TypeError):
            pass

    # Pattern 3: MMDDYYYY or MMDDYY
    match = re.search(r'(\d{2})(\d{2})(\d{4})', filename)
    if match:
        try:
            month, day, year = int(match.group(1)), int(match.group(2)), int(match.group(3))
            return datetime(year, month, day)
        except (ValueError, TypeError):
            pass

    match = re.search(r'(\d{2})(\d{2})(\d{2})', filename)
    if match:
        try:
            month, day, year = int(match.group(1)), int(match.group(2)), int(match.group(3))
            year = 2000 + year if year < 100 else year
            return datetime(year, month, day)
        except (ValueError, TypeError):
            pass

    # Pattern 4: FY_YYYY or FY YYYY (fiscal year — use end of year)
    match = re.search(r'FY[_\s]?(\d{4})', filename, re.IGNORECASE)
    if match:
        try:
            year = int(match.group(1))
            return datetime(year, 12, 31)
        except (ValueError, TypeError):
            pass

    # Pattern 5: Just YYYY (assume end of year)
    match = re.search(r'\b(20\d{2})\b', filename)
    if match:
        try:
            year = int(match.group(1))
            return datetime(year, 12, 31)
        except (ValueError, TypeError):
            pass

    return None


# ==================== RENT ROLL EXTRACTION ====================

def extract_rent_roll(filepath: str) -> Dict[str, Any]:
    """
    Parse a rent roll Excel file and return structured data.

    The extraction must handle the common rent roll format where:
    - Each unit has a primary row with: unit number, unit type, sqft, status, resident, lease dates, market rent
    - Below each unit row are charge detail rows (Rent, Internet, Parking, Valet Trash, etc.)
    - A "Charge Total" row aggregates all charges for that unit
    - A summary row at the bottom has property totals

    Returns:
    {
        "document_date": "2026-01-28",
        "property_name": "1160 Hammond",
        "units": [...],
        "summary": {...}
    }
    """
    try:
        wb = openpyxl.load_workbook(filepath, read_only=False, data_only=True)
        ws = wb.worksheets[0]  # Assume first sheet

        # Step 1: Extract document date from first 5 rows
        document_date = _extract_document_date_from_sheet(ws)

        # Step 2: Extract property name from first 5 rows
        property_name = _extract_property_name_from_sheet(ws)

        # Step 3: Find header row
        header_row_num, headers = _find_rent_roll_header(ws)
        if header_row_num == 0:
            logger.warning("Could not find rent roll header row")
            wb.close()
            return {
                "document_date": None,
                "property_name": property_name,
                "units": [],
                "summary": {}
            }

        # Step 4: Parse units
        units = _parse_rent_roll_units(ws, header_row_num, headers)

        # Step 4b: Safety net — ensure every unit has in_place_rent derived
        # from charge_details. Guards against edge cases where _finalize_unit
        # or the Charge Total handler didn't set the value.
        for unit in units:
            cd = unit.get("charge_details") or {}
            if cd and not unit.get("in_place_rent"):
                base = find_base_rent(cd)
                unit["in_place_rent"] = base if base > 0 else sum(cd.values())

        # Step 5: Calculate summary
        summary = _calculate_rent_roll_summary(units)

        # Step 6: Parse RealPage "Summary Groups" block if present —
        # authoritative aggregates straight from the PMS, used by the
        # upload handler to cross-check the row-level counts.
        realpage_summary = _parse_realpage_summary_groups(ws)

        wb.close()

        return {
            "document_date": document_date.isoformat() if document_date else None,
            "property_name": property_name,
            "units": units,
            "summary": summary,
            "realpage_summary": realpage_summary,
        }

    except Exception as e:
        logger.exception(f"Error extracting rent roll: {e}")
        return {
            "document_date": None,
            "property_name": None,
            "units": [],
            "summary": {},
            "error": str(e)
        }


def _extract_document_date_from_sheet(ws: Worksheet) -> Optional[datetime]:
    """Look for a date in the first 5 rows"""
    for row_idx in range(1, 6):
        for cell in ws[row_idx]:
            if cell.value is None:
                continue

            # Check if cell contains a datetime object
            if isinstance(cell.value, datetime):
                return cell.value

            # Try to parse as date string
            cell_str = str(cell.value).strip()
            # Common patterns
            for fmt in ["%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d", "%m-%d-%Y", "%B %d, %Y"]:
                try:
                    return datetime.strptime(cell_str, fmt)
                except (ValueError, TypeError):
                    continue

    return None


def _extract_property_name_from_sheet(ws: Worksheet) -> Optional[str]:
    """Extract property name from first 5 rows"""
    for row_idx in range(1, 6):
        for cell in ws[row_idx]:
            if cell.value is None:
                continue

            cell_str = str(cell.value).strip()
            # Property names are usually in larger text in the first few rows
            # Look for patterns like "1160 Hammond" or "Property Name"
            if len(cell_str) > 5 and len(cell_str) < 100:
                # Skip if it looks like a date or number
                if not re.match(r'^\d+[/\-]\d+', cell_str) and not cell_str.lower() in ["unit", "status", "resident"]:
                    return cell_str

    return None


def _find_rent_roll_header(ws: Worksheet, max_scan: int = 15) -> Tuple[int, Dict[int, str]]:
    """
    Find the header row in a rent roll.
    Returns (row_number, {col_index: header_name})
    """
    keywords = ["unit", "sqft", "sq ft", "sf", "status", "rent", "resident", "tenant",
                "lease", "move in", "market", "type", "bldg", "scheduled", "charge"]

    best_row = 0
    best_headers = {}
    best_score = 0

    for row_idx in range(1, max_scan + 1):
        row = ws[row_idx]
        headers_dict = {}
        score = 0

        for col_idx, cell in enumerate(row, start=1):
            if cell.value is not None:
                header = str(cell.value).strip()
                headers_dict[col_idx] = header

                # Score this header
                header_lower = header.lower()
                if any(kw in header_lower for kw in keywords):
                    score += 1

        if score > best_score and len(headers_dict) >= 4:  # At least 4 columns
            best_score = score
            best_row = row_idx
            best_headers = headers_dict

    return best_row, best_headers


# ─── Base rent identification ─────────────────────────────────────────────────

BASE_RENT_SYNONYMS = [
    'rent', 'base rent', 'contract rent', 'lease rent', 'monthly rent',
    'unit rent', 'apt rent', 'apartment rent', 'residential rent',
    'scheduled rent', 'gross rent', 'net rent', 'basic rent',
]

EXCLUDE_PREFIXES = [
    'amenity', 'pet', 'parking', 'storage', 'garage', 'concierge',
    'trash', 'internet', 'cable', 'utility', 'month to month', 'valet',
]


def find_base_rent(charge_details: dict) -> float:
    """
    Identify the base rent amount from a charge_details dict.
    Uses a 3-step approach: exact match → fuzzy match → single unambiguous code.
    """
    if not charge_details:
        return 0.0

    # Step 1: exact match on known synonyms
    for code, amount in charge_details.items():
        if code.lower().strip() in BASE_RENT_SYNONYMS:
            return float(amount or 0)

    # Step 2: fuzzy match via rapidfuzz
    codes = list(charge_details.keys())
    best_match = process.extractOne(
        'rent', codes,
        scorer=fuzz.token_sort_ratio,
        score_cutoff=70,
    )
    if best_match:
        matched_code = best_match[0]
        if not any(matched_code.lower().startswith(ex) for ex in EXCLUDE_PREFIXES):
            return float(charge_details[matched_code] or 0)

    # Step 3: single unambiguous rent code
    rent_codes = [
        c for c in charge_details
        if 'rent' in c.lower()
        and not any(ex in c.lower() for ex in EXCLUDE_PREFIXES)
    ]
    if len(rent_codes) == 1:
        return float(charge_details[rent_codes[0]] or 0)

    return 0.0


def _finalize_unit(unit: Dict[str, Any], charge_details: dict) -> None:
    """Finalize a unit by setting charge_details and computing in_place_rent."""
    if charge_details:
        unit["charge_details"] = charge_details.copy()
        base = find_base_rent(charge_details)
        if base > 0:
            unit["in_place_rent"] = base
        elif unit.get("in_place_rent", 0) == 0:
            unit["in_place_rent"] = sum(charge_details.values())
    elif not unit.get("charge_details"):
        unit["charge_details"] = {}


# Hardcoded charge code names for sheets that put charge codes in col A
_COL_A_CHARGE_CODES = {
    "rent", "amenity rent", "internet", "parking", "parking fee",
    "package concierge", "valet trash", "trash", "pet rent", "garage",
    "storage", "utility", "water", "sewer", "cable", "admin fee",
}

# Patterns that identify charge code / description columns — these must NOT
# be claimed by the greedy "charge" / "rent" in_place_rent fallback keywords.
_CHARGE_CODE_PATTERNS = ["charge code", "charge description", "description", "ledger"]


def _parse_rent_roll_units(ws: Worksheet, header_row: int, headers: Dict[int, str]) -> List[Dict[str, Any]]:
    """
    Parse unit rows from rent roll.

    Each unit block consists of:
    - A primary row with unit info (unit number in unit col)
    - Optional charge detail sub-rows (blank unit col, charge code in charge_code col or col A)
    - A "Charge Total" row aggregating all charges
    """
    col_map = _map_rent_roll_columns(headers)

    logger.warning(
        "RENT ROLL PARSE DEBUG: header_row=%d, col_map=%s, headers=%s",
        header_row, col_map, {k: v for k, v in headers.items()}
    )

    units: List[Dict[str, Any]] = []
    current_unit: Optional[Dict[str, Any]] = None
    charge_details: Dict[str, float] = {}

    unit_col = col_map.get("unit")
    charge_code_col = col_map.get("charge_code")
    rent_col = col_map.get("in_place_rent")

    for row_idx in range(header_row + 1, ws.max_row + 1):
        row = ws[row_idx]

        unit_val = _get_cell_value(row, unit_col)
        charge_code_val = _get_cell_value(row, charge_code_col) if charge_code_col else None

        # Scan entire row text for Charge Total / summary markers
        row_text = " ".join(str(c.value or "") for c in row).lower()

        if row_idx <= header_row + 5:
            logger.warning(
                "RENT ROLL ROW %d: unit_val=%r, charge_code_val=%r, row_text=%s",
                row_idx, unit_val, charge_code_val, row_text[:100]
            )

        # ── Charge Total row ──
        if "charge total" in row_text or "total charges" in row_text:
            if current_unit is not None:
                total = _get_numeric_value(row, rent_col)
                current_unit["charge_details"] = charge_details.copy()
                # Prefer base rent from charge_details; fall back to charge total
                base = find_base_rent(charge_details)
                if base > 0:
                    current_unit["in_place_rent"] = base
                elif total:
                    current_unit["in_place_rent"] = total
                elif charge_details:
                    current_unit["in_place_rent"] = sum(charge_details.values())
                charge_details = {}
            continue

        # ── Summary row → stop parsing ──
        # "charge code" catches the "Summary of Charges by Charge Code"
        # banner that precedes the per-code summary block at the end of
        # RealPage rent rolls; without it, short-alpha codes like `aprk`
        # or `mtm` can slip through as fake unit rows.
        if unit_val and any(
            kw in unit_val.lower()
            for kw in ["total", "summary", "grand total", "property", "charge code"]
        ):
            break

        # ── Charge detail sub-row: blank unit col, charge code populated ──
        if not unit_val and charge_code_val:
            amount = _get_numeric_value(row, rent_col)
            logger.warning("CHARGE DETAIL: row=%d, charge_code=%r, amount=%r", row_idx, charge_code_val, amount)
            if amount:
                charge_details[charge_code_val] = amount
            continue

        # ── Charge detail via col A (fallback for sheets without charge_code column) ──
        if unit_val and unit_val.lower().strip() in _COL_A_CHARGE_CODES:
            amount = _get_numeric_value(row, rent_col)
            if amount:
                charge_details[unit_val] = amount
            continue

        # ── Skip rows without a unit value ──
        if not unit_val:
            # Fallback: if charge_code_col is missing, scan for a text cell
            # paired with an amount at rent_col — treat as charge sub-row.
            if current_unit is not None and rent_col:
                amount = _get_numeric_value(row, rent_col)
                if amount:
                    code_name = None
                    for cell in row:
                        v = cell.value
                        if isinstance(v, str) and v.strip():
                            code_name = v.strip()
                            break
                    if code_name:
                        charge_details[code_name] = amount
            continue

        # ── New unit header row ──
        # Finalize previous unit
        if current_unit is not None:
            logger.warning(
                "FINALIZE UNIT: unit=%s, charge_details=%s, in_place_rent=%s",
                current_unit.get('unit_number'), charge_details, current_unit.get('in_place_rent')
            )
            _finalize_unit(current_unit, charge_details)
            units.append(current_unit)
            charge_details = {}

        current_unit = _parse_unit_row(row, col_map)

        # Unit header row may also carry a charge code (e.g. "Amenity Rent" in col L)
        if charge_code_val:
            amount = _get_numeric_value(row, rent_col)
            if amount:
                charge_details[charge_code_val] = amount

    # Finalize last unit
    if current_unit is not None:
        _finalize_unit(current_unit, charge_details)
        units.append(current_unit)

    return units


def _map_rent_roll_columns(headers: Dict[int, str]) -> Dict[str, int]:
    """Map standardized field names to column indices"""
    col_map = {}

    for col_idx, header in headers.items():
        header_lower = header.lower().strip()

        # Unit identification
        if any(kw in header_lower for kw in ["bldg-unit", "unit", "unit #", "unit number", "apt", "apartment"]):
            if "unit" not in col_map:  # Prefer first match
                col_map["unit"] = col_idx

        # Unit type
        if "type" in header_lower and "property" not in header_lower:
            col_map["unit_type"] = col_idx

        # Square footage
        if any(kw in header_lower for kw in ["sqft", "sq ft", "sf", "square"]):
            col_map["sqft"] = col_idx

        # Status
        if "status" in header_lower:
            col_map["status"] = col_idx

        # Resident
        if any(kw in header_lower for kw in ["resident", "tenant", "name"]):
            if "resident" not in col_map:
                col_map["resident"] = col_idx

        # Move in date
        if "move in" in header_lower or "move-in" in header_lower:
            col_map["move_in"] = col_idx

        # Move out date — used to derive occupancy for exports without a Status column
        if "move out" in header_lower or "move-out" in header_lower:
            col_map["move_out"] = col_idx

        # Lease dates
        if "lease" in header_lower:
            if "start" in header_lower or "from" in header_lower:
                col_map["lease_start"] = col_idx
            elif "end" in header_lower or "to" in header_lower or "expire" in header_lower:
                col_map["lease_end"] = col_idx
        # "Lease Expiration" without the word "lease end" — RealPage uses this header
        if "expiration" in header_lower and "lease_end" not in col_map:
            col_map["lease_end"] = col_idx

        # Rents
        if "market" in header_lower and "rent" in header_lower:
            col_map["market_rent"] = col_idx
        elif any(kw in header_lower for kw in [
            "in place", "in-place", "current rent", "actual rent",
            "contract rent", "lease rent", "monthly rent",
            "charge amount", "scheduled rent", "scheduled charges",
            "amount", "charge", "rent",
        ]):
            # Skip columns that are actually charge code/description columns
            # (bare "charge"/"rent" keywords would otherwise greedily match them)
            if not any(kw in header_lower for kw in _CHARGE_CODE_PATTERNS):
                if "in_place_rent" not in col_map:
                    col_map["in_place_rent"] = col_idx

        # Charge code column (used for charge detail sub-rows)
        if any(kw in header_lower for kw in ["charge code", "charge description", "description", "ledger"]):
            if "charge_code" not in col_map:
                col_map["charge_code"] = col_idx

    return col_map


def _parse_unit_row(row, col_map: Dict[str, int]) -> Dict[str, Any]:
    """Parse a single unit row.

    `is_occupied` is left as `None` when the source has no Status column
    and no Move Out date — the downstream normalizer then derives
    occupancy from the date fields and resident_name.
    """
    status_val = _get_cell_value(row, col_map.get("status"))
    move_out_val = _get_date_value(row, col_map.get("move_out"))
    unit = {
        "unit_number": _get_cell_value(row, col_map.get("unit", 1)),
        "unit_type": _get_cell_value(row, col_map.get("unit_type")),
        "sqft": _get_numeric_value(row, col_map.get("sqft")),
        "status": status_val,
        "is_occupied": None,  # derived below or in the normalizer
        "resident_name": _get_cell_value(row, col_map.get("resident")),
        "move_in_date": _get_date_value(row, col_map.get("move_in")),
        "move_out_date": move_out_val,
        "lease_start": _get_date_value(row, col_map.get("lease_start")),
        "lease_end": _get_date_value(row, col_map.get("lease_end")),
        "market_rent": _get_numeric_value(row, col_map.get("market_rent")),
        "in_place_rent": 0,  # Will be set from charge details or charge total
        "charge_details": {}
    }

    status_lower = str(status_val).lower() if status_val else ""
    if status_lower:
        if any(kw in status_lower for kw in ["vacant", "unrented", "model", "employee", "down", "admin"]):
            unit["is_occupied"] = False
        elif any(kw in status_lower for kw in ["occupied", "current", "notice"]):
            unit["is_occupied"] = True
    # Leave is_occupied=None if status is missing or unmatched — the
    # normalizer will derive from move_out_date / lease_end / resident_name.

    return unit


def _get_cell_value(row, col_idx: Optional[int]) -> Optional[str]:
    """Get cell value as string"""
    if col_idx is None or col_idx < 1 or col_idx > len(row):
        return None
    cell = row[col_idx - 1]  # row is 0-indexed
    if cell.value is None:
        return None
    return str(cell.value).strip()


def _get_numeric_value(row, col_idx: Optional[int]) -> Optional[float]:
    """Get cell value as number"""
    if col_idx is None or col_idx < 1 or col_idx > len(row):
        return None
    cell = row[col_idx - 1]
    if cell.value is None:
        return None

    # Try direct conversion
    try:
        return float(cell.value)
    except (ValueError, TypeError):
        pass

    # Try parsing string
    try:
        val_str = str(cell.value).replace(",", "").replace("$", "").strip()
        return float(val_str)
    except (ValueError, TypeError):
        return None


def _get_date_value(row, col_idx: Optional[int]) -> Optional[datetime]:
    """Get cell value as datetime"""
    if col_idx is None or col_idx < 1 or col_idx > len(row):
        return None
    cell = row[col_idx - 1]
    if cell.value is None:
        return None

    # If already datetime
    if isinstance(cell.value, datetime):
        return cell.value

    # Try parsing common date formats
    val_str = str(cell.value).strip()
    for fmt in ["%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d", "%m-%d-%Y"]:
        try:
            return datetime.strptime(val_str, fmt)
        except (ValueError, TypeError):
            continue

    return None


_SUMMARY_LABEL_KEYS = {
    "current/notice/vacant residents": "current_notice_vacant",
    "future residents/applicants": "future_applicants",
    "occupied units": "occupied",
    "total non rev units": "non_rev",
    "non rev units": "non_rev",
    "total vacant units": "vacant",
    "totals:": "totals",
    "grand total": "totals",
}


def _parse_realpage_summary_groups(ws: Worksheet) -> Optional[Dict[str, Any]]:
    """Locate and parse the RealPage 'Summary Groups' block.

    The block sits near the foot of the sheet and carries the authoritative
    aggregates straight from RealPage: total units, occupied, vacant,
    non-rev, future applicants, total sqft / market rent / lease charges,
    and physical occupancy %. Returns None when the block is absent.
    """
    header_row_idx: Optional[int] = None
    for row_idx in range(1, ws.max_row + 1):
        for cell in ws[row_idx]:
            if isinstance(cell.value, str) and "summary groups" in cell.value.strip().lower():
                header_row_idx = row_idx
                break
        if header_row_idx is not None:
            break

    if header_row_idx is None:
        return None

    # RealPage wraps long labels like "# Of Units" and "Square Footage"
    # across two physical rows. Merge the header row with the row below
    # by column index so the label match sees the full string.
    header_rows = [ws[header_row_idx]]
    if header_row_idx + 1 <= ws.max_row:
        header_rows.append(ws[header_row_idx + 1])
    merged: Dict[int, str] = {}
    for hrow in header_rows:
        for cell in hrow:
            if not isinstance(cell.value, str):
                continue
            piece = cell.value.strip()
            if not piece:
                continue
            prior = merged.get(cell.column, "")
            merged[cell.column] = f"{prior} {piece}".strip() if prior else piece

    col_idx: Dict[str, int] = {}
    for col, label_raw in merged.items():
        label = label_raw.lower()
        if "square footage" in label or label in ("sqft", "sq ft"):
            col_idx["sqft"] = col
        elif "market rent" in label:
            col_idx["market_rent"] = col
        elif "lease charges" in label or "actual rent" in label:
            col_idx["lease_charges"] = col
        elif (
            "# of units" in label
            or "# units" in label
            or label in ("units", "unit count")
        ):
            col_idx["units"] = col
        elif "% occupied" in label or "occupancy" in label or "% occ" in label:
            col_idx["occupancy_pct"] = col

    if "units" not in col_idx:
        logger.warning(
            "RealPage Summary Groups at row %d found, but could not map '# Units' column; skipping",
            header_row_idx,
        )
        return None

    # Skip over the continuation header row when one was consumed.
    data_start = header_row_idx + len(header_rows)
    rows_by_key: Dict[str, Dict[str, Any]] = {}
    for row_idx in range(data_start, min(data_start + 30, ws.max_row + 1)):
        row = ws[row_idx]

        label = None
        for cell in row:
            if isinstance(cell.value, str) and cell.value.strip():
                label = cell.value.strip().lower()
                break
        if not label:
            continue

        # Stop at the charge-code block.
        if "summary of charges" in label or "charge code" in label:
            break

        key = None
        for pat, mapped in _SUMMARY_LABEL_KEYS.items():
            if pat in label:
                key = mapped
                break
        if key is None:
            continue

        def _num(col: Optional[int]) -> Optional[float]:
            if col is None:
                return None
            cell = row[col - 1] if col - 1 < len(row) else None
            v = cell.value if cell is not None else None
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                return float(v)
            if isinstance(v, str):
                s = v.strip().replace(",", "").replace("$", "").replace("%", "")
                try:
                    return float(s)
                except ValueError:
                    return None
            return None

        rows_by_key[key] = {
            "sqft": _num(col_idx.get("sqft")),
            "market_rent": _num(col_idx.get("market_rent")),
            "lease_charges": _num(col_idx.get("lease_charges")),
            "units": _num(col_idx.get("units")),
            "occupancy_pct": _num(col_idx.get("occupancy_pct")),
        }

    if not rows_by_key:
        return None

    totals = rows_by_key.get("totals") or rows_by_key.get("current_notice_vacant")
    occupied = rows_by_key.get("occupied", {}) or {}
    vacant = rows_by_key.get("vacant", {}) or {}
    non_rev = rows_by_key.get("non_rev", {}) or {}
    future = rows_by_key.get("future_applicants", {}) or {}

    def _as_int(v: Any) -> Optional[int]:
        if v is None:
            return None
        try:
            return int(round(float(v)))
        except (TypeError, ValueError):
            return None

    return {
        "source_row": header_row_idx,
        "total_units": _as_int((totals or {}).get("units")),
        "occupied_units": _as_int(occupied.get("units")),
        "vacant_units": _as_int(vacant.get("units")),
        "non_rev_units": _as_int(non_rev.get("units")),
        "future_leases": _as_int(future.get("units")),
        "total_sqft": _as_int((totals or {}).get("sqft")),
        "total_market_rent": (totals or {}).get("market_rent"),
        "total_lease_charges": (totals or {}).get("lease_charges"),
        "physical_occupancy_pct": (
            occupied.get("occupancy_pct")
            or (rows_by_key.get("current_notice_vacant", {}) or {}).get("occupancy_pct")
        ),
    }


def _calculate_rent_roll_summary(units: List[Dict[str, Any]]) -> Dict[str, Any]:
    """Calculate summary statistics for rent roll"""
    if not units:
        return {}

    # Filter out non-revenue units (model, employee)
    revenue_units = [u for u in units if u.get("status") and
                     not any(kw in str(u["status"]).lower() for kw in ["model", "employee"])]

    total_units = len(revenue_units)
    occupied_units = sum(1 for u in revenue_units if u.get("is_occupied"))
    vacant_units = total_units - occupied_units

    # Calculate averages (only for occupied units with data)
    market_rents = [u["market_rent"] for u in revenue_units if u.get("market_rent")]
    in_place_rents = [u["in_place_rent"] for u in revenue_units if u.get("is_occupied") and u.get("in_place_rent")]
    sqfts = [u["sqft"] for u in revenue_units if u.get("sqft")]

    avg_market_rent = sum(market_rents) / len(market_rents) if market_rents else None
    avg_in_place_rent = sum(in_place_rents) / len(in_place_rents) if in_place_rents else None
    avg_sqft = sum(sqfts) / len(sqfts) if sqfts else None

    # Physical occupancy
    physical_occupancy_pct = (occupied_units / total_units * 100) if total_units > 0 else 0

    # Loss to lease
    loss_to_lease_pct = None
    if avg_market_rent and avg_in_place_rent and avg_market_rent > 0:
        loss_to_lease_pct = ((avg_market_rent - avg_in_place_rent) / avg_market_rent) * 100

    return {
        "total_units": total_units,
        "occupied_units": occupied_units,
        "vacant_units": vacant_units,
        "physical_occupancy_pct": round(physical_occupancy_pct, 2) if physical_occupancy_pct else None,
        "avg_market_rent": round(avg_market_rent, 2) if avg_market_rent else None,
        "avg_in_place_rent": round(avg_in_place_rent, 2) if avg_in_place_rent else None,
        "avg_sqft": round(avg_sqft, 2) if avg_sqft else None,
        "loss_to_lease_pct": round(loss_to_lease_pct, 2) if loss_to_lease_pct else None
    }


# ==================== T-12 EXTRACTION ====================

def extract_t12(filepath: str) -> Dict[str, Any]:
    """
    Parse a T-12 operating statement Excel file and return structured data.

    Returns:
    {
        "fiscal_year": 2025,
        "property_name": "1160 Hammond",
        "summary": {...},
        "monthly": {...},
        "line_items": {...}
    }
    """
    try:
        wb = openpyxl.load_workbook(filepath, read_only=False, data_only=True)
        ws = wb.worksheets[0]

        # Step 1: Find month header row and year
        month_row, month_cols, fiscal_year = _find_t12_month_headers(ws)
        if month_row == 0:
            logger.warning("Could not find T-12 month headers")
            wb.close()
            return {
                "fiscal_year": None,
                "property_name": None,
                "summary": {},
                "monthly": {},
                "line_items": {}
            }

        # Step 2: Find "Total Year" column
        total_col = _find_total_year_column(ws, month_row)

        # Step 2b: Detect layout — find the label column adaptively
        layout = detect_t12_layout(ws)
        label_col = layout.get("label_col")
        if label_col is not None:
            logger.info("T12 layout detection: label_col=%d, has_gl_codes=%s",
                        label_col, layout.get("has_gl_codes"))

        # Step 3: Extract property name
        property_name = _extract_property_name_from_sheet(ws)

        # Step 4: Parse all line items (using detected label column)
        line_items = _parse_t12_line_items(ws, month_row, month_cols, total_col,
                                           label_col=label_col)

        # Step 5: Extract summary values
        summary = _extract_t12_summary(line_items, total_col)

        # Step 6: Extract monthly data for key metrics
        monthly = _extract_t12_monthly(line_items, month_cols)

        # Calculate ratios
        if summary.get("total_revenue") and summary.get("total_revenue") > 0:
            summary["expense_ratio_pct"] = round((summary.get("total_operating_expenses", 0) / summary["total_revenue"]) * 100, 2)
            summary["noi_margin_pct"] = round((summary.get("net_operating_income", 0) / summary["total_revenue"]) * 100, 2)

        wb.close()

        return {
            "fiscal_year": fiscal_year,
            "property_name": property_name,
            "summary": summary,
            "monthly": monthly,
            "line_items": line_items,
            "expense_ratio_pct": summary.get("expense_ratio_pct"),
            "noi_margin_pct": summary.get("noi_margin_pct")
        }

    except Exception as e:
        logger.exception(f"Error extracting T-12: {e}")
        return {
            "fiscal_year": None,
            "property_name": None,
            "summary": {},
            "monthly": {},
            "line_items": {},
            "error": str(e)
        }


def _find_t12_month_headers(ws: Worksheet, max_scan: int = 20) -> Tuple[int, Dict[str, int], Optional[int]]:
    """
    Find the row with month headers (Jan, Feb, ...).
    Returns (row_number, {month_name: col_index}, fiscal_year)

    Handles month headers as:
    - Text strings: "Jan", "January", "Jan-25", "Jan 2025"
    - Date values: datetime(2025, 1, 1) (common in Yardi/RealPage exports)
    - Numeric month references: "1/2025", "01/25"
    """
    month_names = ["january", "february", "march", "april", "may", "june",
                   "july", "august", "september", "october", "november", "december"]
    month_abbr = ["jan", "feb", "mar", "apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec"]

    for row_idx in range(1, max_scan + 1):
        row = ws[row_idx]
        month_cols = {}
        fiscal_year = None

        # Check if this row contains month names
        for col_idx, cell in enumerate(row, start=1):
            if cell.value is None:
                continue

            matched = False

            # Handle datetime objects (Excel date values for month headers)
            if isinstance(cell.value, datetime):
                month_idx = cell.value.month - 1  # 0-indexed
                month_cols[month_abbr[month_idx].capitalize()] = col_idx
                # Track fiscal year from the date values
                if fiscal_year is None or cell.value.year > fiscal_year:
                    fiscal_year = cell.value.year
                matched = True

            if not matched:
                cell_str = str(cell.value).lower().strip()

                # Check for month names or abbreviations
                for i, month_name in enumerate(month_names):
                    if month_name in cell_str or month_abbr[i] in cell_str:
                        month_cols[month_abbr[i].capitalize()] = col_idx
                        # Try to extract year from the same cell (e.g., "Jan-25", "Jan 2025")
                        year_match = re.search(r'(20\d{2})', cell_str)
                        if year_match:
                            fiscal_year = int(year_match.group(1))
                        elif re.search(r'[\-/](\d{2})$', cell_str):
                            # Handle 2-digit year like "Jan-25"
                            yr2 = re.search(r'[\-/](\d{2})$', cell_str)
                            if yr2:
                                yr = int(yr2.group(1))
                                fiscal_year = 2000 + yr if yr < 100 else yr
                        break

        # If we found at least 6 months, this is probably the header row
        if len(month_cols) >= 6:
            # Try to find fiscal year from the row above (if not already found from dates)
            if fiscal_year is None and row_idx > 1:
                prev_row = ws[row_idx - 1]
                for cell in prev_row:
                    if cell.value and isinstance(cell.value, (int, float)):
                        year = int(cell.value)
                        if 2000 <= year <= 2100:
                            fiscal_year = year
                            break
                    elif cell.value:
                        # Try parsing year from string
                        match = re.search(r'(20\d{2})', str(cell.value))
                        if match:
                            fiscal_year = int(match.group(1))
                            break

            logger.info("T12 month headers found at row %d with %d months, fiscal_year=%s",
                        row_idx, len(month_cols), fiscal_year)
            return row_idx, month_cols, fiscal_year

    logger.warning("T12 month headers NOT found after scanning %d rows", max_scan)
    return 0, {}, None


def _find_total_year_column(ws: Worksheet, month_row: int) -> Optional[int]:
    """Find the 'Total Year' or 'Annual' column"""
    row = ws[month_row]

    for col_idx, cell in enumerate(row, start=1):
        if cell.value is None:
            continue

        cell_str = str(cell.value).lower().strip()
        if any(kw in cell_str for kw in ["total", "annual", "year", "ytd"]):
            return col_idx

    # If not found, assume it's the last column with data
    max_col = 0
    for col_idx, cell in enumerate(row, start=1):
        if cell.value is not None:
            max_col = col_idx

    return max_col if max_col > 0 else None


def _parse_t12_line_items(ws: Worksheet, month_row: int, month_cols: Dict[str, int],
                          total_col: Optional[int],
                          label_col: Optional[int] = None) -> Dict[str, Dict[str, Any]]:
    """
    Parse all line items from T-12 with monthly values.
    Returns: {"Line Item Name": {"Jan": 123, "Feb": 456, ..., "Total": 7890}}

    label_col: 0-indexed column for line item labels (auto-detected or column A fallback)
    """
    line_items = {}
    label_col_idx = label_col if label_col is not None else 0

    for row_idx in range(month_row + 1, ws.max_row + 1):
        row = ws[row_idx]

        # Get line item name from detected label column
        if label_col_idx >= len(row):
            continue
        item_name_cell = row[label_col_idx]
        if item_name_cell.value is None:
            continue

        item_name = str(item_name_cell.value).strip()

        # Skip empty or separator rows
        if not item_name or len(item_name) < 2:
            continue

        # Strip GL codes (like "403011") from the beginning
        item_name = re.sub(r'^\d{5,}\s*', '', item_name).strip()

        # Get monthly values
        monthly_values = {}
        for month_abbr, col_idx in month_cols.items():
            value = _get_numeric_value(row, col_idx)
            if value is not None:
                monthly_values[month_abbr] = value

        # Get total
        total_value = None
        if total_col:
            total_value = _get_numeric_value(row, total_col)

        # Store line item
        if monthly_values or total_value is not None:
            line_items[item_name] = {
                **monthly_values,
                "Total": total_value
            }

    return line_items


def _extract_t12_summary(line_items: Dict[str, Dict[str, Any]], total_col: Optional[int]) -> Dict[str, Any]:
    """
    Extract key summary values from line items using fuzzy matching.

    Uses the T12 taxonomy and match_line_item() for robust matching
    against real-world line item label variations.
    """
    summary = {}

    # Skip words — lines containing these are likely false positives
    skip_words = ["excl ", "excluding", "net potential", "after "]

    # Match each line item using fuzzy matching
    for item_name, values in line_items.items():
        item_lower = item_name.lower()

        # Skip false positives
        if any(sw in item_lower for sw in skip_words):
            continue

        # Use fuzzy matching to find the canonical taxonomy key
        taxonomy_key = match_line_item(item_name, T12_TAXONOMY)
        if taxonomy_key is None:
            continue

        # Map taxonomy key to summary field name
        summary_field = TAXONOMY_TO_SUMMARY_FIELD.get(taxonomy_key)
        if summary_field is None:
            continue

        # Only set if not already set (first match wins — more specific items
        # appear earlier in spreadsheets)
        if summary_field not in summary and values.get("Total") is not None:
            summary[summary_field] = values["Total"]
            logger.debug("Fuzzy matched '%s' → %s = %s", item_name, summary_field, values["Total"])

    # Fallback: if NOI is missing, calculate from total_revenue - total_operating_expenses
    if summary.get("net_operating_income") is None:
        rev = summary.get("total_revenue")
        exp = summary.get("total_operating_expenses")
        if rev is not None and exp is not None:
            summary["net_operating_income"] = rev - abs(exp)
            logger.info("T12 NOI computed as fallback: revenue(%s) - expenses(%s) = %s",
                        rev, exp, summary["net_operating_income"])

    # Fallback: if total_revenue is missing but GPR and other components exist
    if summary.get("total_revenue") is None and summary.get("gross_potential_rent") is not None:
        gpr = summary["gross_potential_rent"]
        deductions = sum(abs(summary.get(k, 0) or 0) for k in [
            "loss_to_lease", "vacancy_loss", "concessions", "bad_debt"
        ])
        other = summary.get("other_income", 0) or 0
        summary["total_revenue"] = gpr - deductions + other
        logger.info("T12 total_revenue computed as fallback: GPR(%s) - deductions(%s) + other(%s) = %s",
                     gpr, deductions, other, summary["total_revenue"])

    logger.info("T12 summary extraction (fuzzy): %s", {k: v for k, v in summary.items() if v is not None})
    return summary


def _extract_t12_monthly(line_items: Dict[str, Dict[str, Any]], month_cols: Dict[str, int]) -> Dict[str, Dict[str, float]]:
    """Extract monthly data for NOI, Revenue, and Expenses using fuzzy matching"""
    monthly = {
        "noi": {},
        "revenue": {},
        "expenses": {}
    }

    # Map taxonomy keys to monthly dict keys
    monthly_targets = {
        "noi": "noi",
        "egi": "revenue",
        "total_opex": "expenses",
    }

    # Find NOI, Revenue, and Expenses line items via fuzzy matching
    for item_name, values in line_items.items():
        taxonomy_key = match_line_item(item_name, T12_TAXONOMY)
        if taxonomy_key not in monthly_targets:
            continue

        target = monthly_targets[taxonomy_key]
        if monthly[target]:
            continue  # Already found this one

        for month in month_cols.keys():
            if month in values:
                monthly[target][month] = values[month]

    logger.info("T12 monthly extraction: noi=%d months, revenue=%d months, expenses=%d months",
                len(monthly["noi"]), len(monthly["revenue"]), len(monthly["expenses"]))
    return monthly


# ==================== TIER 2: CLAUDE AI FALLBACK ====================

def _worksheet_to_text(filepath: str) -> str:
    """Convert all sheets to a text representation for Claude."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    output = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        output.append(f"\n=== Sheet: {sheet_name} ===\n")
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            if any(c.strip() for c in cells):
                output.append("\t".join(cells))
    wb.close()
    return "\n".join(output[:500])  # Cap at 500 lines to manage token cost


async def extract_t12_with_ai_fallback(filepath: str) -> dict:
    """
    Try fuzzy matching first, fall back to Claude if it fails.
    Tier 1: Fuzzy keyword matching (fast, no API cost)
    Tier 2: Claude AI extraction (handles truly unusual formats)
    """
    # Tier 1: Fuzzy matching
    result = extract_t12(filepath)

    # Check if extraction actually produced data
    summary = result.get("summary", {})
    key_fields = ['gross_potential_rent', 'net_operating_income',
                  'total_operating_expenses', 'total_revenue']
    populated = sum(1 for f in key_fields if summary.get(f) not in [None, 0, 0.0])

    if populated >= 2:
        logger.info("T12 fuzzy extraction succeeded: %d/4 key fields populated", populated)
        return result

    logger.warning("T12 fuzzy extraction weak (%d/4 fields). Falling back to Claude AI.", populated)

    # Tier 2: Send to Claude
    try:
        import anthropic

        sheet_text = _worksheet_to_text(filepath)

        client = anthropic.Anthropic(
            api_key=settings.ANTHROPIC_API_KEY,
            base_url="https://api.anthropic.com",
            timeout=120.0,
        )
        response = client.messages.create(
            model="claude-sonnet-4-5-20250929",
            max_tokens=4096,
            system="You are a CRE financial data extraction expert. Extract T12 trailing 12-month financial data from this spreadsheet content. Return ONLY valid JSON.",
            messages=[{
                "role": "user",
                "content": f"""Extract the T12 (trailing 12 months) financial summary from this spreadsheet data.

Return JSON with these fields (use null if not found, use annual/total figures not monthly):
{{
  "gross_potential_rent": <gross scheduled/potential rent, annual>,
  "vacancy_loss": <vacancy loss, annual, as positive number>,
  "concessions": <concessions, annual>,
  "bad_debt": <bad debt/credit loss, annual>,
  "other_income": <other/ancillary income, annual>,
  "total_revenue": <effective gross income, annual>,
  "total_operating_expenses": <total operating expenses, annual>,
  "net_operating_income": <net operating income, annual>,
  "real_estate_taxes": <real estate taxes, annual>,
  "insurance": <insurance, annual>,
  "management_fee": <management fee, annual>,
  "repairs_maintenance": <repairs & maintenance, annual>,
  "utilities": <utilities, annual>,
  "payroll": <payroll/salaries, annual>,
  "fiscal_year": <the year of the T12 period, e.g. 2025>,
  "period_end_month": <last month of the trailing period, e.g. "December">
}}

Spreadsheet content:
{sheet_text}"""
            }],
        )

        # Parse Claude's JSON response
        ai_text = response.content[0].text
        # Strip markdown code fences if present
        ai_text = re.sub(r'^```(?:json)?\s*', '', ai_text.strip())
        ai_text = re.sub(r'\s*```$', '', ai_text.strip())

        ai_result = json.loads(ai_text)
        logger.info("Claude AI T12 extraction succeeded: NOI=%s", ai_result.get('net_operating_income'))

        # Merge AI results into the standard result structure
        result["summary"] = ai_result
        if ai_result.get("fiscal_year"):
            result["fiscal_year"] = ai_result["fiscal_year"]

        return result

    except Exception as e:
        logger.error("Claude AI T12 extraction failed: %s", e)
        return result  # Return the weak Tier 1 result as last resort


# ==================== DETAILED T12 EXTRACTION (for T12 Mapper) ====================

def _parse_t12_line_items_detailed(
    ws: Worksheet,
    month_row: int,
    month_cols: Dict[str, int],
    total_col: Optional[int],
    label_col: Optional[int] = None,
) -> List[Dict[str, Any]]:
    """
    Parse all T12 line items preserving row order, section headers, subtotals, and GL codes.

    Returns a list of dicts (one per row), each with:
      raw_label, gl_code, row_index, section, subsection,
      is_subtotal, is_section_header, monthly_values, annual_total
    """
    MONTH_ORDER = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
                   "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

    label_col_idx = label_col if label_col is not None else 0
    items: List[Dict[str, Any]] = []

    current_section = "revenue"  # default until we see an expense header
    current_subsection: Optional[str] = None

    # Patterns for detecting section boundaries
    expense_patterns = [
        r'^expenses?\b', r'^operating\s+expenses?\b', r'^total\s+expenses?\b',
        r'^controllable\b', r'^property\s+operating',
    ]
    revenue_patterns = [
        r'^income\b', r'^revenue\b', r'^rental\s+income', r'^total\s+income',
    ]
    subtotal_patterns = [
        r'^total\b', r'^subtotal\b', r'^net\s+operating', r'^noi\b',
        r'^effective\s+gross', r'^gross\s+potential',
    ]
    header_keywords = [
        "income", "revenue", "expenses", "operating expenses",
        "payroll & benefits", "payroll and benefits",
        "repairs & maintenance", "repairs and maintenance",
        "contract services", "administrative", "general & administrative",
        "marketing & leasing", "utilities", "other income",
        "rental income", "non-controllable", "controllable",
    ]

    for row_idx in range(month_row + 1, ws.max_row + 1):
        row = ws[row_idx]

        if label_col_idx >= len(row):
            continue
        cell_value = row[label_col_idx].value
        if cell_value is None:
            continue

        raw_label = str(cell_value).strip()
        if not raw_label or len(raw_label) < 2:
            continue

        # Extract GL code before stripping
        gl_code = None
        gl_match = re.match(r'^(\d{4,})\s*[\-:]?\s*', raw_label)
        if gl_match:
            gl_code = gl_match.group(1)

        # Clean label (strip GL codes)
        clean_label = re.sub(r'^\d{5,}\s*[\-:]?\s*', '', raw_label).strip()
        if not clean_label:
            clean_label = raw_label

        # Get monthly values
        monthly_values: Dict[str, float] = {}
        for month_abbr, col_idx in month_cols.items():
            value = _get_numeric_value(row, col_idx)
            if value is not None:
                monthly_values[month_abbr] = value

        # Get total
        annual_total = None
        if total_col:
            annual_total = _get_numeric_value(row, total_col)

        # Determine if this is a section header (no numeric values, matches header keywords)
        has_values = bool(monthly_values) or annual_total is not None
        clean_lower = clean_label.lower().strip()
        is_section_header = False
        is_subtotal = False

        if not has_values:
            # Likely a section header or subsection header
            if any(clean_lower == kw or clean_lower.startswith(kw) for kw in header_keywords):
                is_section_header = True
                # Update subsection tracking
                current_subsection = clean_label
                # Check if this switches us to expenses section
                for pat in expense_patterns:
                    if re.match(pat, clean_lower):
                        current_section = "expense"
                        current_subsection = None
                        break
        else:
            # Check if subtotal
            for pat in subtotal_patterns:
                if re.match(pat, clean_lower):
                    is_subtotal = True
                    break

            # Also check for section transition at "Total Income" or expense headers
            if re.match(r'^total\s+income', clean_lower) or re.match(r'^total\s+revenue', clean_lower):
                is_subtotal = True
            # Check for expense section start even if values present
            for pat in expense_patterns:
                if re.match(pat, clean_lower):
                    current_section = "expense"
                    is_section_header = True
                    break

        # Compute T1/T2/T3 from monthly values
        t1_value = None
        t2_value = None
        t3_value = None
        if monthly_values:
            # Get months in order, take the last ones
            ordered_months = [m for m in MONTH_ORDER if m in monthly_values]
            if ordered_months:
                t1_value = monthly_values.get(ordered_months[-1])
                if len(ordered_months) >= 2:
                    t2_value = sum(monthly_values[m] for m in ordered_months[-2:])
                if len(ordered_months) >= 3:
                    t3_sum = sum(monthly_values[m] for m in ordered_months[-3:])
                    t3_value = t3_sum * 4  # annualized

        items.append({
            "raw_label": clean_label,
            "gl_code": gl_code,
            "row_index": row_idx,
            "section": current_section,
            "subsection": current_subsection if not is_section_header else None,
            "is_subtotal": is_subtotal,
            "is_section_header": is_section_header,
            "monthly_values": monthly_values if monthly_values else None,
            "annual_total": annual_total,
            "t1_value": t1_value,
            "t2_value": t2_value,
            "t3_value": t3_value,
        })

    return items


def _auto_categorize_line_items(
    items: List[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    """
    Auto-categorize each line item using match_line_item() and map to WDIS categories.
    Adds mapped_category and auto_confidence to each item dict.
    """
    from app.schemas.t12_mapping import TAXONOMY_KEY_TO_CATEGORY

    for item in items:
        if item["is_section_header"] or item["is_subtotal"]:
            item["mapped_category"] = None
            item["auto_confidence"] = None
            continue

        raw = item["raw_label"]
        cleaned = raw.strip().lower()
        cleaned = re.sub(r'^[\d\-\.]+\s*', '', cleaned)
        cleaned = re.sub(r'\s+', ' ', cleaned)

        # Use existing fuzzy matcher
        taxonomy_key = match_line_item(raw, T12_TAXONOMY, threshold=50)

        if taxonomy_key:
            category = TAXONOMY_KEY_TO_CATEGORY.get(taxonomy_key)
            if category:
                # Compute confidence from fuzzy score
                all_keywords = []
                for key, entry in T12_TAXONOMY.items():
                    if key == taxonomy_key:
                        for kw in entry["keywords"]:
                            all_keywords.append(kw)
                        break

                best_score = 0
                if all_keywords:
                    match_result = process.extractOne(cleaned, all_keywords, scorer=fuzz.token_sort_ratio)
                    if match_result:
                        best_score = match_result[1] / 100.0  # normalize to 0-1

                item["mapped_category"] = category
                item["auto_confidence"] = round(best_score, 3)
            else:
                # Taxonomy key maps to a subtotal category — skip
                item["mapped_category"] = None
                item["auto_confidence"] = None
        else:
            # No match — leave unmapped
            item["mapped_category"] = None
            item["auto_confidence"] = None

    return items


def extract_t12_detailed(filepath: str) -> Dict[str, Any]:
    """
    Extended T12 extraction that preserves individual line items with row order,
    section headers, subtotals, GL codes, and auto-categorization.

    Returns the standard extract_t12() result plus a 'detailed_items' list.
    """
    # Run standard extraction first
    result = extract_t12(filepath)

    try:
        wb = openpyxl.load_workbook(filepath, read_only=False, data_only=True)
        ws = wb.worksheets[0]

        month_row, month_cols, fiscal_year = _find_t12_month_headers(ws)
        if month_row == 0:
            wb.close()
            result["detailed_items"] = []
            return result

        total_col = _find_total_year_column(ws, month_row)
        layout = detect_t12_layout(ws)
        label_col = layout.get("label_col")

        # Parse detailed line items
        items = _parse_t12_line_items_detailed(ws, month_row, month_cols, total_col, label_col)

        # Auto-categorize
        items = _auto_categorize_line_items(items)

        wb.close()
        result["detailed_items"] = items

    except Exception as e:
        logger.exception("Error in extract_t12_detailed: %s", e)
        result["detailed_items"] = []

    return result
