"""
Data Bank Extraction Service — Two-phase extraction for Excel uploads

Phase 1: Python structural parse (openpyxl) → detect type, find headers, extract raw data
Phase 2: Claude API semantic extraction → column classification, value normalization

Supports:
- Sales comp spreadsheets
- Pipeline/supply tracker spreadsheets
- Underwriting models
"""
import json
import logging
import re
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
from sqlalchemy.orm import Session

from app.config import settings
from app.models.data_bank import DataBankDocument, PipelineProject, SalesComp
from app.services.claude_extraction_service import parse_json_response

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# CRE vocabulary for document type detection
# ---------------------------------------------------------------------------

SALES_COMP_VOCAB = {
    "sheet": ["comp", "comparable", "sales", "transaction", "trade", "closed", "sale"],
    "header": [
        "property_name", "property name", "sale_price", "sale price", "cap_rate",
        "cap rate", "price_per_unit", "price per unit", "price/unit", "buyer",
        "seller", "sale_date", "sale date", "close date", "closing date",
        "price_per_sf", "price per sf", "$/sf", "$/unit", "occupancy",
        "year_built", "year built", "vintage", "units", "unit count",
    ],
}

PIPELINE_VOCAB = {
    "sheet": ["pipeline", "supply", "construction", "delivery", "development",
              "new supply", "under construction", "proposed"],
    "header": [
        "project_name", "project name", "developer", "delivery", "status",
        "units", "unit count", "start", "completion", "lease_up", "lease up",
        "under construction", "proposed", "quarter", "delivery date",
    ],
}

UNDERWRITING_VOCAB = {
    "sheet": ["summary", "proforma", "pro forma", "cash flow", "returns",
              "assumptions", "rent roll", "underwriting", "model", "output"],
    "header": [
        "noi", "irr", "cap rate", "dscr", "ltv", "equity multiple",
        "cash on cash", "reversion", "exit cap", "hold period",
        "gross revenue", "vacancy", "opex", "net operating income",
    ],
}


# ---------------------------------------------------------------------------
# Phase 0: Document type detection
# ---------------------------------------------------------------------------

def detect_document_type(filepath: str) -> str:
    """
    Detect whether an Excel file contains sales comps, a pipeline tracker,
    or an underwriting model by examining sheet names and header rows.

    Returns: "sales_comps" | "pipeline_tracker" | "underwriting_model" | "unknown"
    """
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    except Exception as e:
        logger.warning(f"Could not open workbook for type detection: {e}")
        return "unknown"

    scores = {"sales_comps": 0, "pipeline_tracker": 0, "underwriting_model": 0}

    try:
        # Score by sheet names
        sheet_names_lower = [s.lower() for s in wb.sheetnames]

        for sn in sheet_names_lower:
            for kw in SALES_COMP_VOCAB["sheet"]:
                if kw in sn:
                    scores["sales_comps"] += 3
            for kw in PIPELINE_VOCAB["sheet"]:
                if kw in sn:
                    scores["pipeline_tracker"] += 3
            for kw in UNDERWRITING_VOCAB["sheet"]:
                if kw in sn:
                    scores["underwriting_model"] += 3

        # Score by header content in first 10 rows of each sheet
        for ws in wb.worksheets:
            headers_text = []
            for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
                for cell in row:
                    if cell is not None:
                        headers_text.append(str(cell).lower().strip())

            header_blob = " ".join(headers_text)

            for kw in SALES_COMP_VOCAB["header"]:
                if kw in header_blob:
                    scores["sales_comps"] += 2

            for kw in PIPELINE_VOCAB["header"]:
                if kw in header_blob:
                    scores["pipeline_tracker"] += 2

            for kw in UNDERWRITING_VOCAB["header"]:
                if kw in header_blob:
                    scores["underwriting_model"] += 2

    finally:
        wb.close()

    # Pick the highest score
    best = max(scores, key=scores.get)  # type: ignore[arg-type]
    if scores[best] == 0:
        return "unknown"

    logger.info(f"Document type scores: {scores} → {best}")
    return best


# ---------------------------------------------------------------------------
# Phase 1: Structural parse — read raw tabular data from Excel
# ---------------------------------------------------------------------------

def _find_best_sheet(wb: openpyxl.Workbook, vocab_keywords: List[str]) -> Any:
    """Pick the sheet that best matches the given keywords, or the one with most data rows."""
    best_sheet = None
    best_score = -1

    for ws in wb.worksheets:
        score = 0
        name_lower = ws.title.lower()
        for kw in vocab_keywords:
            if kw in name_lower:
                score += 10

        # Count data rows (rough heuristic)
        row_count = 0
        for row in ws.iter_rows(min_row=1, max_row=500, values_only=True):
            if any(c is not None for c in row):
                row_count += 1
        score += row_count

        if score > best_score:
            best_score = score
            best_sheet = ws

    return best_sheet


def _find_header_row(ws: Any, keywords: List[str], max_scan: int = 10) -> Tuple[int, List[str]]:
    """
    Scan the first `max_scan` rows for the header row that contains the most
    CRE-relevant keywords. Returns (row_number_1indexed, list_of_header_values).
    """
    best_row = 1
    best_headers: List[str] = []
    best_score = -1

    for row_idx in range(1, max_scan + 1):
        cells = []
        for cell in ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True):
            cells = [str(c).strip() if c is not None else "" for c in cell]
            break

        if not cells:
            continue

        blob = " ".join(c.lower() for c in cells)
        score = sum(1 for kw in keywords if kw in blob)

        if score > best_score:
            best_score = score
            best_row = row_idx
            best_headers = cells

    # Handle split header rows (merge row above if it adds context)
    if best_row > 1:
        row_above = []
        for cell in ws.iter_rows(min_row=best_row - 1, max_row=best_row - 1, values_only=True):
            row_above = [str(c).strip() if c is not None else "" for c in cell]
            break

        merged_headers = []
        for i, h in enumerate(best_headers):
            above = row_above[i] if i < len(row_above) else ""
            if above and not h:
                merged_headers.append(above)
            elif above and h and above.lower() != h.lower():
                merged_headers.append(f"{above} {h}")
            else:
                merged_headers.append(h)

        # Only use merged if it adds information
        merged_blob = " ".join(c.lower() for c in merged_headers)
        merged_score = sum(1 for kw in keywords if kw in merged_blob)
        if merged_score > best_score:
            best_headers = merged_headers

    return best_row, best_headers


def _extract_raw_rows(
    ws: Any, header_row: int, headers: List[str], max_rows: int = 2000
) -> List[Dict[str, Any]]:
    """Read data rows starting after the header row, skipping empty rows."""
    rows = []
    for row in ws.iter_rows(min_row=header_row + 1, max_row=header_row + max_rows, values_only=True):
        values = list(row)
        # Skip completely empty rows
        if not any(v is not None and str(v).strip() != "" for v in values):
            continue

        record = {}
        for i, h in enumerate(headers):
            if h and i < len(values):
                record[h] = values[i]

        rows.append(record)

    return rows


# ---------------------------------------------------------------------------
# Phase 2: Claude API — column classification and value normalization
# ---------------------------------------------------------------------------

CLASSIFICATION_PROMPT_COMPS = """You are a CRE data analyst. I have an Excel spreadsheet with sales comp data.

Here are the column headers:
{headers}

Here are the first {sample_count} sample data rows:
{sample_rows}

Map each column header to the correct CRE field. Return a JSON object with:
{{
  "column_mapping": {{
    "Original Header Name": "mapped_field_name"
  }}
}}

Valid mapped_field_name values:
- property_name, market, metro, submarket, county, state, address
- property_type, sale_date, year_built, year_renovated, units, avg_unit_sf
- avg_eff_rent, sale_price, price_per_unit, price_per_sf
- cap_rate, cap_rate_qualifier, occupancy
- buyer, seller, notes
- SKIP (for columns that don't map to any field)

Rules:
- "Year Built" columns that contain "1987/2017" format → map to "year_built" (we'll parse renovation separately)
- "Cap Rate" → "cap_rate"
- Map location fields precisely: "MSA" or "Metro" → metro, "Market" → market, "Submarket" → submarket
- "Sale Price" or "Purchase Price" → sale_price
- "PPU" or "Price Per Unit" or "$/Unit" → price_per_unit

Return ONLY the JSON object, no other text."""

CLASSIFICATION_PROMPT_PIPELINE = """You are a CRE data analyst. I have an Excel spreadsheet with pipeline/supply tracker data.

Here are the column headers:
{headers}

Here are the first {sample_count} sample data rows:
{sample_rows}

Map each column header to the correct CRE field. Return a JSON object with:
{{
  "column_mapping": {{
    "Original Header Name": "mapped_field_name"
  }}
}}

Valid mapped_field_name values:
- project_name, address, county, metro, submarket
- units, status, developer, delivery_quarter, start_quarter, property_type
- SKIP (for columns that don't map to any field)

Rules:
- "Project Name" or "Property Name" or "Development" → project_name
- "Status" values like "Under Construction", "Lease-Up", "Proposed" will be normalized later
- "Delivery" or "Completion" or "Expected Delivery" → delivery_quarter
- "Start" or "Construction Start" → start_quarter

Return ONLY the JSON object, no other text."""

NORMALIZATION_PROMPT_COMPS = """You are a CRE data normalization engine. Clean and normalize these sales comp records.

Column mapping (original → field): {column_mapping}

Raw records to normalize:
{records}

Return a JSON array of normalized records. For each record:
{{
  "property_name": "string or null",
  "market": "string or null",
  "metro": "string or null",
  "submarket": "string or null",
  "county": "string or null",
  "state": "string or null (2-letter code)",
  "address": "string or null",
  "property_type": "string or null",
  "sale_date": "YYYY-MM-DD or null",
  "year_built": integer or null,
  "year_renovated": integer or null,
  "units": integer or null,
  "avg_unit_sf": number or null,
  "avg_eff_rent": number or null,
  "sale_price": number or null (in dollars, NOT millions),
  "price_per_unit": number or null,
  "price_per_sf": number or null,
  "cap_rate": number or null (as decimal, e.g. 0.055 NOT 5.5),
  "cap_rate_qualifier": "string or null",
  "occupancy": number or null (as decimal, e.g. 0.95),
  "buyer": "string or null",
  "seller": "string or null",
  "notes": "string or null"
}}

Normalization rules:
- Sale price in millions ("$63.88M" or "63.88") → multiply by 1,000,000 → 63880000
- Cap rate as percentage (5.5, 5.5%) → divide by 100 → 0.055
- Cap rate already decimal (0.055) → keep as-is (if value < 0.3, it's already decimal)
- Occupancy as percentage (95, 95%) → divide by 100 → 0.95
- Year built with renovation ("1987/2017") → year_built=1987, year_renovated=2017
- TBD / N/A / "-" / empty / "n/a" → null
- Remove $ signs, commas, % signs from numbers
- State names → 2-letter code (Texas → TX)

Return ONLY the JSON array, no other text."""

NORMALIZATION_PROMPT_PIPELINE = """You are a CRE data normalization engine. Clean and normalize these pipeline project records.

Column mapping (original → field): {column_mapping}

Raw records to normalize:
{records}

Return a JSON array of normalized records. For each record:
{{
  "project_name": "string or null",
  "address": "string or null",
  "county": "string or null",
  "metro": "string or null",
  "submarket": "string or null",
  "units": integer or null,
  "status": "lease_up" | "under_construction" | "proposed" | null,
  "developer": "string or null",
  "delivery_quarter": "Q1 2026 format or null",
  "start_quarter": "Q1 2025 format or null",
  "property_type": "string or null"
}}

Status normalization:
- "Lease-Up", "Lease Up", "Leasing", "Stabilizing" → "lease_up"
- "Under Construction", "UC", "In Construction", "Building" → "under_construction"
- "Proposed", "Planned", "Approved", "Pre-Development", "Planning" → "proposed"
- Unknown / empty → null

Rules:
- TBD / N/A / "-" / empty → null
- Delivery dates: normalize to "Q1 2026" format
- Remove extra whitespace

Return ONLY the JSON array, no other text."""

UNDERWRITING_EXTRACTION_PROMPT = """You are a CRE underwriting analyst. Extract key metrics from this underwriting model data.

Sheet names in the workbook: {sheet_names}

Data from key sheets (first 80 rows each):
{sheet_data}

Extract and return a JSON object with:
{{
  "model_type": "underwriting" | "proforma" | "rent_roll" | "other",
  "property_info": {{
    "property_name": "string or null",
    "address": "string or null",
    "property_type": "string or null",
    "units": integer or null,
    "total_sf": number or null
  }},
  "assumptions": {{
    "purchase_price": number or null,
    "cap_rate_going_in": number or null (decimal),
    "exit_cap_rate": number or null (decimal),
    "hold_period_years": integer or null,
    "rent_growth_rate": number or null (decimal),
    "expense_growth_rate": number or null (decimal),
    "vacancy_rate": number or null (decimal),
    "ltv": number or null (decimal),
    "interest_rate": number or null (decimal)
  }},
  "returns": {{
    "unlevered_irr": number or null (decimal),
    "levered_irr": number or null (decimal),
    "equity_multiple": number or null,
    "avg_cash_on_cash": number or null (decimal),
    "noi_year1": number or null,
    "noi_stabilized": number or null
  }},
  "rent_roll_summary": {{
    "avg_in_place_rent": number or null,
    "avg_market_rent": number or null,
    "occupancy": number or null (decimal)
  }},
  "warnings": ["list of any data quality issues"]
}}

Rules:
- Cap rates, IRRs, occupancy as decimals (5.5% → 0.055)
- Prices in actual dollars, not millions
- If a field is not found, use null
- Add warnings for anything unusual

Return ONLY the JSON object, no other text."""


def _get_claude_client():
    """Initialize Anthropic client. Returns None if API key not configured."""
    try:
        import anthropic
    except ImportError:
        logger.warning("anthropic package not installed — falling back to Python-only extraction")
        return None

    if not settings.ANTHROPIC_API_KEY or settings.ANTHROPIC_API_KEY == "your-api-key-here":
        logger.warning("ANTHROPIC_API_KEY not configured — falling back to Python-only extraction")
        return None

    return anthropic.Anthropic(
        api_key=settings.ANTHROPIC_API_KEY,
        base_url="https://api.anthropic.com",
    )


def _call_claude(client: Any, prompt: str, max_tokens: int = 4096) -> str:
    """Send a prompt to Claude and return the response text."""
    message = client.messages.create(
        model="claude-sonnet-4-5-20250929",
        max_tokens=max_tokens,
        messages=[{"role": "user", "content": prompt}],
    )
    return message.content[0].text


def _classify_columns(
    client: Any, headers: List[str], sample_rows: List[Dict], doc_type: str
) -> Dict[str, str]:
    """Use Claude to classify spreadsheet columns. Falls back to fuzzy matching if no client."""
    if client is None:
        return _fallback_column_mapping(headers, doc_type)

    prompt_template = (
        CLASSIFICATION_PROMPT_COMPS if doc_type == "sales_comps"
        else CLASSIFICATION_PROMPT_PIPELINE
    )

    # Limit sample rows for context size
    samples = sample_rows[:5]
    prompt = prompt_template.format(
        headers=json.dumps(headers),
        sample_count=len(samples),
        sample_rows=json.dumps(samples, default=str),
    )

    try:
        response_text = _call_claude(client, prompt)
        result = parse_json_response(response_text)
        return result.get("column_mapping", {})
    except Exception as e:
        logger.warning(f"Claude column classification failed: {e}, using fallback")
        return _fallback_column_mapping(headers, doc_type)


def _normalize_records(
    client: Any,
    records: List[Dict],
    column_mapping: Dict[str, str],
    doc_type: str,
    batch_size: int = 25,
) -> List[Dict]:
    """Use Claude to normalize records in batches. Falls back to basic Python normalization."""
    if client is None:
        return _fallback_normalize(records, column_mapping, doc_type)

    prompt_template = (
        NORMALIZATION_PROMPT_COMPS if doc_type == "sales_comps"
        else NORMALIZATION_PROMPT_PIPELINE
    )

    all_normalized: List[Dict] = []

    for i in range(0, len(records), batch_size):
        batch = records[i:i + batch_size]
        prompt = prompt_template.format(
            column_mapping=json.dumps(column_mapping),
            records=json.dumps(batch, default=str),
        )

        try:
            response_text = _call_claude(client, prompt)
            parsed = parse_json_response(response_text)

            # Response should be a list
            if isinstance(parsed, list):
                all_normalized.extend(parsed)
            elif isinstance(parsed, dict) and "records" in parsed:
                all_normalized.extend(parsed["records"])
            else:
                logger.warning(f"Unexpected normalization response format, batch {i}")
                all_normalized.extend(_fallback_normalize(batch, column_mapping, doc_type))
        except Exception as e:
            logger.warning(f"Claude normalization failed for batch {i}: {e}, using fallback")
            all_normalized.extend(_fallback_normalize(batch, column_mapping, doc_type))

    return all_normalized


# ---------------------------------------------------------------------------
# Fallback extraction (Python-only, no Claude)
# ---------------------------------------------------------------------------

_COMP_HEADER_MAP = {
    "property name": "property_name", "property": "property_name", "name": "property_name",
    "market": "market", "metro": "metro", "msa": "metro",
    "submarket": "submarket", "sub market": "submarket", "sub-market": "submarket",
    "county": "county", "state": "state",
    "address": "address", "location": "address",
    "property type": "property_type", "type": "property_type", "asset type": "property_type",
    "sale date": "sale_date", "close date": "sale_date", "closing date": "sale_date",
    "date": "sale_date",
    "year built": "year_built", "vintage": "year_built", "yr built": "year_built",
    "year renovated": "year_renovated", "renovation": "year_renovated",
    "units": "units", "unit count": "units", "# units": "units", "total units": "units",
    "avg unit sf": "avg_unit_sf", "avg sf": "avg_unit_sf", "unit sf": "avg_unit_sf",
    "avg eff rent": "avg_eff_rent", "eff rent": "avg_eff_rent", "avg rent": "avg_eff_rent",
    "rent": "avg_eff_rent",
    "sale price": "sale_price", "purchase price": "sale_price", "price": "sale_price",
    "price per unit": "price_per_unit", "ppu": "price_per_unit", "$/unit": "price_per_unit",
    "price/unit": "price_per_unit",
    "price per sf": "price_per_sf", "$/sf": "price_per_sf", "price/sf": "price_per_sf",
    "cap rate": "cap_rate", "cap": "cap_rate",
    "cap rate qualifier": "cap_rate_qualifier",
    "occupancy": "occupancy", "occ": "occupancy", "occ%": "occupancy",
    "buyer": "buyer", "purchaser": "buyer",
    "seller": "seller", "vendor": "seller",
    "notes": "notes", "comments": "notes",
}

_PIPELINE_HEADER_MAP = {
    "project name": "project_name", "property name": "project_name",
    "project": "project_name", "name": "project_name", "development": "project_name",
    "address": "address", "location": "address",
    "county": "county", "metro": "metro", "msa": "metro",
    "submarket": "submarket",
    "units": "units", "unit count": "units", "# units": "units", "total units": "units",
    "status": "status", "phase": "status",
    "developer": "developer", "owner": "developer",
    "delivery": "delivery_quarter", "delivery date": "delivery_quarter",
    "completion": "delivery_quarter", "expected delivery": "delivery_quarter",
    "start": "start_quarter", "construction start": "start_quarter",
    "start date": "start_quarter",
    "property type": "property_type", "type": "property_type",
}


def _fallback_column_mapping(headers: List[str], doc_type: str) -> Dict[str, str]:
    """Fuzzy match headers to known CRE field names.

    Uses longest-key-wins strategy for substring matching so that specific
    keys like ``price per sf`` beat generic keys like ``price``.
    """
    header_map = _COMP_HEADER_MAP if doc_type == "sales_comps" else _PIPELINE_HEADER_MAP
    mapping = {}

    for h in headers:
        if not h:
            continue
        h_lower = h.lower().strip()

        # Exact match
        if h_lower in header_map:
            mapping[h] = header_map[h_lower]
            continue

        # Substring match — pick the longest (most specific) matching key
        best_field: Optional[str] = None
        best_len = 0
        for key, field in header_map.items():
            if key in h_lower or h_lower in key:
                if len(key) > best_len:
                    best_len = len(key)
                    best_field = field

        if best_field is not None:
            mapping[h] = best_field
        else:
            mapping[h] = "SKIP"

    return mapping


def _clean_numeric(val: Any) -> Optional[float]:
    """Parse a numeric value from potentially messy input."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)

    s = str(val).strip()
    if not s or s.lower() in ("tbd", "n/a", "-", "na", "none", "null", ""):
        return None

    # Check for millions notation
    millions = False
    if s.upper().endswith("M"):
        millions = True
        s = s[:-1]

    # Remove $, commas, % signs
    s = s.replace("$", "").replace(",", "").replace("%", "").strip()

    try:
        num = float(s)
        if millions:
            num *= 1_000_000
        return num
    except ValueError:
        return None


def _clean_int(val: Any) -> Optional[int]:
    """Parse an integer value."""
    num = _clean_numeric(val)
    return int(num) if num is not None else None


def _parse_year_built(val: Any) -> Tuple[Optional[int], Optional[int]]:
    """Parse year built, handling '1987/2017' renovation format."""
    if val is None:
        return None, None

    s = str(val).strip()
    if not s or s.lower() in ("tbd", "n/a", "-"):
        return None, None

    # Check for year/renovation format
    match = re.match(r"(\d{4})\s*[/\\-]\s*(\d{4})", s)
    if match:
        return int(match.group(1)), int(match.group(2))

    num = _clean_int(val)
    if num and 1800 <= num <= 2100:
        return num, None
    return None, None


def _normalize_cap_rate(val: Any) -> Optional[float]:
    """Normalize cap rate to decimal (0.055 not 5.5)."""
    num = _clean_numeric(val)
    if num is None:
        return None
    # If > 0.3, it's a percentage — convert to decimal
    if num > 0.3:
        return num / 100.0
    return num


def _normalize_occupancy(val: Any) -> Optional[float]:
    """Normalize occupancy to decimal (0.95 not 95)."""
    num = _clean_numeric(val)
    if num is None:
        return None
    if num > 1.5:
        return num / 100.0
    return num


def _normalize_status(val: Any) -> Optional[str]:
    """Normalize pipeline status to enum values."""
    if val is None:
        return None

    s = str(val).lower().strip()
    if not s or s in ("tbd", "n/a", "-"):
        return None

    if any(kw in s for kw in ("lease", "leasing", "stabiliz")):
        return "lease_up"
    if any(kw in s for kw in ("under construction", "uc", "in construction", "building")):
        return "under_construction"
    if any(kw in s for kw in ("proposed", "planned", "approved", "pre-dev", "planning")):
        return "proposed"

    return None


def _parse_sale_date(val: Any) -> Optional[str]:
    """Parse a sale date into YYYY-MM-DD string."""
    if val is None:
        return None

    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")

    s = str(val).strip()
    if not s or s.lower() in ("tbd", "n/a", "-"):
        return None

    # Try common date formats
    for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%m-%d-%Y", "%m/%d/%y", "%b %Y", "%B %Y", "%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue

    return None


def _fallback_normalize(
    records: List[Dict], column_mapping: Dict[str, str], doc_type: str
) -> List[Dict]:
    """Normalize records using Python only (no Claude)."""
    normalized = []

    for record in records:
        mapped: Dict[str, Any] = {}

        for orig_header, value in record.items():
            field = column_mapping.get(orig_header, "SKIP")
            if field == "SKIP":
                continue
            mapped[field] = value

        if doc_type == "sales_comps":
            year_built, year_renovated = _parse_year_built(mapped.get("year_built"))
            row = {
                "property_name": str(mapped.get("property_name", "")).strip() or None,
                "market": str(mapped.get("market", "")).strip() or None,
                "metro": str(mapped.get("metro", "")).strip() or None,
                "submarket": str(mapped.get("submarket", "")).strip() or None,
                "county": str(mapped.get("county", "")).strip() or None,
                "state": str(mapped.get("state", "")).strip() or None,
                "address": str(mapped.get("address", "")).strip() or None,
                "property_type": str(mapped.get("property_type", "")).strip() or None,
                "sale_date": _parse_sale_date(mapped.get("sale_date")),
                "year_built": year_built,
                "year_renovated": year_renovated if year_renovated else _clean_int(mapped.get("year_renovated")),
                "units": _clean_int(mapped.get("units")),
                "avg_unit_sf": _clean_numeric(mapped.get("avg_unit_sf")),
                "avg_eff_rent": _clean_numeric(mapped.get("avg_eff_rent")),
                "sale_price": _clean_numeric(mapped.get("sale_price")),
                "price_per_unit": _clean_numeric(mapped.get("price_per_unit")),
                "price_per_sf": _clean_numeric(mapped.get("price_per_sf")),
                "cap_rate": _normalize_cap_rate(mapped.get("cap_rate")),
                "cap_rate_qualifier": str(mapped.get("cap_rate_qualifier", "")).strip() or None,
                "occupancy": _normalize_occupancy(mapped.get("occupancy")),
                "buyer": str(mapped.get("buyer", "")).strip() or None,
                "seller": str(mapped.get("seller", "")).strip() or None,
                "notes": str(mapped.get("notes", "")).strip() or None,
            }
            normalized.append(row)

        elif doc_type == "pipeline_tracker":
            row = {
                "project_name": str(mapped.get("project_name", "")).strip() or None,
                "address": str(mapped.get("address", "")).strip() or None,
                "county": str(mapped.get("county", "")).strip() or None,
                "metro": str(mapped.get("metro", "")).strip() or None,
                "submarket": str(mapped.get("submarket", "")).strip() or None,
                "units": _clean_int(mapped.get("units")),
                "status": _normalize_status(mapped.get("status")),
                "developer": str(mapped.get("developer", "")).strip() or None,
                "delivery_quarter": str(mapped.get("delivery_quarter", "")).strip() or None,
                "start_quarter": str(mapped.get("start_quarter", "")).strip() or None,
                "property_type": str(mapped.get("property_type", "")).strip() or None,
            }
            normalized.append(row)

    return normalized


# ---------------------------------------------------------------------------
# Post-normalization validation and cross-field derivation
# ---------------------------------------------------------------------------

# Reasonable default for average unit SF when not provided
_DEFAULT_AVG_UNIT_SF = 900.0


def _validate_and_derive_comp_fields(record: Dict[str, Any]) -> Tuple[Dict[str, Any], List[str]]:
    """
    Validate and fix cross-field consistency for a single sales comp record.

    Detects the common extraction error where price_per_sf values end up in
    the sale_price column.  Derives missing fields where possible.

    Returns the (possibly-modified) record and a list of warning strings.
    """
    warnings: List[str] = []
    sale_price = record.get("sale_price")
    units = record.get("units")
    price_per_unit = record.get("price_per_unit")
    price_per_sf = record.get("price_per_sf")
    avg_unit_sf = record.get("avg_unit_sf")
    prop_name = record.get("property_name") or "unknown"

    # ------------------------------------------------------------------
    # 1. Detect misplaced price_per_sf in sale_price field
    #    Real multifamily sale_price should be > $1M.
    #    Price/SF is typically $100–$600.
    # ------------------------------------------------------------------
    if sale_price is not None and units is not None and units > 10:
        if sale_price < 10_000:
            warnings.append(
                f"{prop_name}: sale_price={sale_price} looks like $/SF — correcting"
            )
            # Preserve the value as price_per_sf if not already set
            if price_per_sf is None:
                record["price_per_sf"] = sale_price
                price_per_sf = sale_price

            # Derive total sale price from $/SF × unit SF × units
            effective_sf = avg_unit_sf if avg_unit_sf and avg_unit_sf > 0 else _DEFAULT_AVG_UNIT_SF
            record["sale_price"] = round(sale_price * effective_sf * units, 2)
            sale_price = record["sale_price"]

    # ------------------------------------------------------------------
    # 2. Derive price_per_unit if missing or nonsensical
    # ------------------------------------------------------------------
    if (price_per_unit is None or price_per_unit < 1_000) and sale_price and units and units > 0 and sale_price > 100_000:
        record["price_per_unit"] = round(sale_price / units, 2)
        price_per_unit = record["price_per_unit"]

    # ------------------------------------------------------------------
    # 3. Derive price_per_sf if missing
    # ------------------------------------------------------------------
    if price_per_sf is None and sale_price and units and avg_unit_sf:
        total_sf = units * avg_unit_sf
        if total_sf > 0:
            record["price_per_sf"] = round(sale_price / total_sf, 2)

    # ------------------------------------------------------------------
    # 4. Sanity-check final values
    # ------------------------------------------------------------------
    if sale_price is not None and sale_price > 0 and units and units > 10:
        ppu = record.get("price_per_unit")
        if ppu is not None and (ppu < 20_000 or ppu > 2_000_000):
            warnings.append(
                f"{prop_name}: derived price_per_unit={ppu} outside expected range $20K–$2M"
            )

    return record, warnings


def _validate_all_comp_records(records: List[Dict[str, Any]]) -> Tuple[List[Dict[str, Any]], List[str]]:
    """Run validation/derivation on every normalised comp record."""
    all_warnings: List[str] = []
    out: List[Dict[str, Any]] = []
    for rec in records:
        fixed, w = _validate_and_derive_comp_fields(rec)
        all_warnings.extend(w)
        out.append(fixed)
    return out, all_warnings


# ---------------------------------------------------------------------------
# Phase 3+4: Full extraction pipelines — extract, classify, normalize, store
# ---------------------------------------------------------------------------

def extract_sales_comps(
    filepath: str, user_id: str, db: Session, document_id: int
) -> Tuple[List[SalesComp], List[str]]:
    """
    Full extraction pipeline for sales comp spreadsheets.

    Returns: (list of SalesComp records, list of warnings)
    """
    warnings: List[str] = []
    client = _get_claude_client()

    # Phase 1: Structural parse
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    try:
        ws = _find_best_sheet(wb, SALES_COMP_VOCAB["sheet"])
        if ws is None:
            warnings.append("Could not find a suitable sheet for sales comps")
            return [], warnings

        header_row, headers = _find_header_row(ws, SALES_COMP_VOCAB["header"])
        if not headers or all(h == "" for h in headers):
            warnings.append("Could not detect header row")
            return [], warnings

        raw_rows = _extract_raw_rows(ws, header_row, headers)
        if not raw_rows:
            warnings.append("No data rows found after header")
            return [], warnings

        logger.info(f"Extracted {len(raw_rows)} raw rows from sheet '{ws.title}', header row {header_row}")
    finally:
        wb.close()

    # Phase 2: Column classification
    column_mapping = _classify_columns(client, headers, raw_rows[:5], "sales_comps")
    if client is None:
        warnings.append("Claude API unavailable — used fallback column mapping")

    # Phase 3: Normalization
    normalized = _normalize_records(client, raw_rows, column_mapping, "sales_comps")
    if client is None:
        warnings.append("Claude API unavailable — used fallback normalization")

    # Phase 3b: Validate and derive cross-field values
    normalized, validation_warnings = _validate_all_comp_records(normalized)
    warnings.extend(validation_warnings)

    # Phase 4: Store in database
    comp_records: List[SalesComp] = []
    for record in normalized:
        # Parse sale_date string to datetime
        sale_date = None
        if record.get("sale_date"):
            try:
                sale_date = datetime.strptime(record["sale_date"], "%Y-%m-%d")
            except (ValueError, TypeError):
                pass

        comp = SalesComp(
            document_id=document_id,
            user_id=user_id,
            property_name=record.get("property_name"),
            market=record.get("market"),
            metro=record.get("metro"),
            submarket=record.get("submarket"),
            county=record.get("county"),
            state=record.get("state"),
            address=record.get("address"),
            property_type=record.get("property_type"),
            sale_date=sale_date,
            year_built=record.get("year_built"),
            year_renovated=record.get("year_renovated"),
            units=record.get("units"),
            avg_unit_sf=record.get("avg_unit_sf"),
            avg_eff_rent=record.get("avg_eff_rent"),
            sale_price=record.get("sale_price"),
            price_per_unit=record.get("price_per_unit"),
            price_per_sf=record.get("price_per_sf"),
            cap_rate=record.get("cap_rate"),
            cap_rate_qualifier=record.get("cap_rate_qualifier"),
            occupancy=record.get("occupancy"),
            buyer=record.get("buyer"),
            seller=record.get("seller"),
            notes=record.get("notes"),
        )
        db.add(comp)
        comp_records.append(comp)

    db.flush()
    return comp_records, warnings


def extract_pipeline_tracker(
    filepath: str, user_id: str, db: Session, document_id: int
) -> Tuple[List[PipelineProject], List[str]]:
    """
    Full extraction pipeline for pipeline/supply tracker spreadsheets.

    Returns: (list of PipelineProject records, list of warnings)
    """
    warnings: List[str] = []
    client = _get_claude_client()

    # Phase 1: Structural parse
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    try:
        ws = _find_best_sheet(wb, PIPELINE_VOCAB["sheet"])
        if ws is None:
            warnings.append("Could not find a suitable sheet for pipeline data")
            return [], warnings

        header_row, headers = _find_header_row(ws, PIPELINE_VOCAB["header"])
        if not headers or all(h == "" for h in headers):
            warnings.append("Could not detect header row")
            return [], warnings

        raw_rows = _extract_raw_rows(ws, header_row, headers)
        if not raw_rows:
            warnings.append("No data rows found after header")
            return [], warnings

        logger.info(f"Extracted {len(raw_rows)} raw rows from sheet '{ws.title}', header row {header_row}")
    finally:
        wb.close()

    # Phase 2: Column classification
    column_mapping = _classify_columns(client, headers, raw_rows[:5], "pipeline_tracker")
    if client is None:
        warnings.append("Claude API unavailable — used fallback column mapping")

    # Phase 3: Normalization
    normalized = _normalize_records(client, raw_rows, column_mapping, "pipeline_tracker")
    if client is None:
        warnings.append("Claude API unavailable — used fallback normalization")

    # Phase 4: Store in database
    project_records: List[PipelineProject] = []
    for record in normalized:
        project = PipelineProject(
            document_id=document_id,
            user_id=user_id,
            project_name=record.get("project_name"),
            address=record.get("address"),
            county=record.get("county"),
            metro=record.get("metro"),
            submarket=record.get("submarket"),
            units=record.get("units"),
            status=record.get("status"),
            developer=record.get("developer"),
            delivery_quarter=record.get("delivery_quarter"),
            start_quarter=record.get("start_quarter"),
            property_type=record.get("property_type"),
        )
        db.add(project)
        project_records.append(project)

    db.flush()
    return project_records, warnings


def extract_underwriting_model(
    filepath: str, user_id: str, db: Session, document_id: int
) -> Tuple[Dict[str, Any], List[str]]:
    """
    Extract key metrics from an underwriting model Excel file.

    Returns: (extraction_data dict, list of warnings)
    """
    warnings: List[str] = []
    client = _get_claude_client()

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    try:
        sheet_names = wb.sheetnames

        # Read first 80 rows from the most relevant sheets
        sheet_data_parts: List[str] = []
        target_sheets = []

        for ws in wb.worksheets:
            name_lower = ws.title.lower()
            is_relevant = any(kw in name_lower for kw in UNDERWRITING_VOCAB["sheet"])
            if is_relevant or len(target_sheets) < 3:
                target_sheets.append(ws)

        for ws in target_sheets[:5]:
            rows_text = [f"\n=== Sheet: {ws.title} ==="]
            for row in ws.iter_rows(min_row=1, max_row=80, values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                if any(c for c in cells):
                    rows_text.append("\t".join(cells))

            sheet_data_parts.append("\n".join(rows_text))

    finally:
        wb.close()

    all_sheet_data = "\n".join(sheet_data_parts)

    if client is None:
        warnings.append("Claude API unavailable — underwriting extraction requires Claude")
        return {"model_type": "underwriting", "warnings": warnings}, warnings

    # Send to Claude for extraction
    prompt = UNDERWRITING_EXTRACTION_PROMPT.format(
        sheet_names=json.dumps(sheet_names),
        sheet_data=all_sheet_data[:80000],  # Keep within context limits
    )

    try:
        response_text = _call_claude(client, prompt)
        extraction_data = parse_json_response(response_text)
        return extraction_data, warnings
    except Exception as e:
        warnings.append(f"Claude underwriting extraction failed: {str(e)}")
        return {"model_type": "underwriting", "warnings": warnings}, warnings


# ---------------------------------------------------------------------------
# Main entry point — orchestrates full extraction for any document type
# ---------------------------------------------------------------------------

def process_data_bank_upload(
    filepath: str,
    user_id: str,
    db: Session,
    document: DataBankDocument,
) -> Tuple[str, int, List[str]]:
    """
    Process a Data Bank Excel upload end-to-end.

    1. Detect document type
    2. Run appropriate extractor
    3. Update document record with results

    Returns: (document_type, record_count, warnings)
    """
    # Detect type
    doc_type = detect_document_type(filepath)
    document.document_type = doc_type
    document.extraction_status = "processing"
    db.flush()

    warnings: List[str] = []
    record_count = 0

    try:
        if doc_type == "sales_comps":
            records, w = extract_sales_comps(filepath, user_id, db, document.id)
            warnings.extend(w)
            record_count = len(records)

        elif doc_type == "pipeline_tracker":
            records, w = extract_pipeline_tracker(filepath, user_id, db, document.id)
            warnings.extend(w)
            record_count = len(records)

        elif doc_type == "underwriting_model":
            extraction_data, w = extract_underwriting_model(filepath, user_id, db, document.id)
            warnings.extend(w)
            document.extraction_data = json.dumps(extraction_data)
            record_count = 1

        else:
            warnings.append(f"Unrecognized document type: {doc_type}")
            document.extraction_status = "failed"
            document.extraction_data = json.dumps({"error": "Unrecognized document type", "warnings": warnings})
            db.commit()
            return doc_type, 0, warnings

        # Success
        document.extraction_status = "completed"
        document.record_count = record_count
        if not document.extraction_data:
            document.extraction_data = json.dumps({"warnings": warnings})
        db.commit()

    except Exception as e:
        logger.exception(f"Extraction failed for document {document.id}")
        db.rollback()
        document.extraction_status = "failed"
        document.extraction_data = json.dumps({"error": str(e), "warnings": warnings})
        db.commit()
        warnings.append(f"Extraction failed: {str(e)}")
        return doc_type, 0, warnings

    return doc_type, record_count, warnings
