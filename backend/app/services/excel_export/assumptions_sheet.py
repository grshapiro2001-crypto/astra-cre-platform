"""Assumptions sheet — all inputs as named ranges, other sheets reference here."""

from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

from app.schemas.underwriting import UWInputs, ScenarioResult

from .styles import put

SHEET = "Assumptions"


def _name(wb, key: str, col: str, row: int) -> None:
    wb.defined_names[key] = DefinedName(
        name=key, attr_text=f"'{SHEET}'!${col}${row}",
    )


def build(wb, inputs: UWInputs, scenario: ScenarioResult, scenario_key: str) -> dict:
    ws = wb.create_sheet(SHEET)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 34
    ws.column_dimensions['B'].width = 20
    for col in range(3, 15):
        ws.column_dimensions[get_column_letter(col)].width = 14

    put(ws, 1, 1, f"Assumptions — {scenario_key.title()} Scenario", style='title')
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    r = 3

    put(ws, r, 1, "Acquisition & Financing", style='section'); r += 1
    pairs = [
        ("PurchasePrice", "Purchase Price", scenario.valuation_summary.purchase_price, 'input_money'),
        ("LoanAmount", "Loan Amount", scenario.debt.loan_amount, 'input_money'),
        ("Equity", "Equity", scenario.debt.equity, 'input_money'),
        ("IntRate", "Interest Rate", inputs.interest_rate, 'input_pct'),
        ("AmortYears", "Amortization (Years)", inputs.amort_years, 'input'),
        ("IOMonths", "Interest-Only (Months)", inputs.io_period_months, 'input'),
        ("LoanTermMonths", "Loan Term (Months)", inputs.loan_term_months, 'input'),
        ("HoldYears", "Hold Period (Years)", inputs.hold_period_years, 'input'),
        ("TerminalCap", "Terminal Cap Rate", getattr(inputs, scenario_key).terminal_cap_rate, 'input_pct'),
        ("SaleCostPct", "Sales Expense %", inputs.sales_expense_pct, 'input_pct'),
        ("MgmtFeePct", "Management Fee %", inputs.mgmt_fee_pct, 'input_pct'),
        ("ReservesPerUnit", "Reserves ($/unit)", inputs.reserves_per_unit, 'input_money'),
        ("TotalUnits", "Total Units", inputs.total_units, 'input'),
    ]
    for key, lbl, val, st in pairs:
        put(ws, r, 1, lbl, style='label')
        put(ws, r, 2, val, style=st)
        _name(wb, key, "B", r)
        r += 1
    r += 1

    put(ws, r, 1, "Growth Rates (Year-over-Year)", style='section'); r += 1
    hold = max(1, inputs.hold_period_years)
    put(ws, r, 1, "Year", style='label_bold')
    for y in range(1, hold + 1):
        put(ws, r, 1 + y, f"Y{y}", style='label_bold')
    r += 1

    curves = [
        ("RentGrowth", "Rental Inflation", inputs.rental_inflation),
        ("ExpenseGrowth", "Expense Inflation", inputs.expense_inflation),
        ("TaxGrowth", "Tax Inflation", inputs.re_tax_inflation),
        ("VacancyPct", "Vacancy %", inputs.vacancy_pct),
        ("ConcessionPct", "Concessions %", inputs.concession_pct),
        ("BadDebtPct", "Bad Debt %", inputs.bad_debt_pct),
    ]
    refs: dict = {"curves": {}}
    for key, lbl, arr in curves:
        put(ws, r, 1, lbl, style='label')
        year_refs = []
        for y in range(hold):
            idx = min(y, len(arr) - 1) if arr else 0
            val = arr[idx] if arr else 0.0
            put(ws, r, 2 + y, val, style='input_pct')
            year_refs.append(f"'{SHEET}'!${get_column_letter(2 + y)}${r}")
        refs["curves"][key] = year_refs
        r += 1
    r += 1

    if inputs.custom_revenue_items or inputs.custom_expense_items:
        put(ws, r, 1, "Custom Line Items", style='section'); r += 1
        for i, h in enumerate(["Label", "Category", "Base Value (Y1)", "Growth Rate", "Start Year"]):
            put(ws, r, 1 + i, h, style='label_bold')
        r += 1
        for item in list(inputs.custom_revenue_items) + list(inputs.custom_expense_items):
            put(ws, r, 1, item.label or item.id, style='label')
            put(ws, r, 2, item.category, style='label')
            put(ws, r, 3, item.base_value, style='money')
            put(ws, r, 4, item.growth_rate, style='percent')
            put(ws, r, 5, item.start_year, style='label')
            r += 1

    ws.freeze_panes = 'B3'
    refs["sheet"] = SHEET
    return refs
