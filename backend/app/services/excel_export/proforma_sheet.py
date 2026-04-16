"""Proforma sheet — operating statement by year with Excel formulas.

Y1 from engine; Y2..YN are `=PrevCell * (1 + Assumptions!GrowthRate)`.
Overrides written as hardcoded values with a comment noting model value.
"""

from openpyxl.utils import get_column_letter

from app.schemas.underwriting import UWInputs, ScenarioResult

from .styles import override, put

SHEET = "Proforma"
REV = [("gpr", "Gross Potential Rent", "RentGrowth"), ("vacancy", "  Vacancy", None),
       ("concessions", "  Concessions", None), ("bad_debt", "  Bad Debt", None),
       ("nru_loss", "  Non-Revenue Units", "RentGrowth"), ("other_income", "Other Income", "RentGrowth")]
EXP = [("controllable_expenses", "Controllable Expenses", "ExpenseGrowth"),
       ("property_taxes", "Property Taxes", "TaxGrowth"),
       ("insurance", "Insurance", "ExpenseGrowth"), ("management_fee", "Management Fee", None)]
DED = {"vacancy": "VacancyPct", "concessions": "ConcessionPct", "bad_debt": "BadDebtPct"}


def _line(ws, r, y, key, curve, years, curves, rm, ti_row):
    cl = get_column_letter(2 + y)
    if key in DED:
        return f"={cl}{rm['gpr']}*{curves.get(DED[key], [None] * (y + 1))[y]}"
    if key == "management_fee":
        return f"={cl}{ti_row}*MgmtFeePct"
    if y == 0:
        return getattr(years[0], key, 0.0) if years else 0.0
    prev = f"{get_column_letter(2 + y - 1)}{r}"
    return f"={prev}*(1+{curves.get(curve, [None] * (y + 1))[y]})"


def build(wb, inputs: UWInputs, scenario: ScenarioResult, scenario_key: str, assumptions_refs: dict) -> dict:
    ws = wb.create_sheet(SHEET)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 32
    years = scenario.dcf.years
    hold = len(years) or max(1, inputs.hold_period_years)
    for c in range(2, 2 + hold):
        ws.column_dimensions[get_column_letter(c)].width = 15

    put(ws, 1, 1, f"Proforma — {scenario_key.title()}", style='label_bold')
    for y in range(1, hold + 1):
        put(ws, 3, 1 + y, f"Year {y}", style='header')

    curves = assumptions_refs.get("curves", {})
    overrides = (inputs.overrides or {}).get(scenario_key, {})
    rm: dict = {}
    r = 4

    put(ws, r, 1, "Revenue", style='section'); r += 1
    for key, lbl, curve in REV:
        put(ws, r, 1, lbl, style='label')
        rm[key] = r
        for y in range(hold):
            cell = put(ws, r, 2 + y, _line(ws, r, y, key, curve, years, curves, rm, 0), style='money')
            ov = f"{key}:{y}"
            if ov in overrides and y < len(years) and years[y].computed_values:
                cell.value = overrides[ov]
                override(cell, years[y].computed_values.get(key, 0.0))
        r += 1

    ti_row = r; rm["total_income"] = r
    put(ws, r, 1, "Total Income (EGI)", style='label_bold')
    for y in range(hold):
        cl = get_column_letter(2 + y)
        put(ws, r, 2 + y, (f"={cl}{rm['gpr']}-{cl}{rm['vacancy']}-{cl}{rm['concessions']}"
                          f"-{cl}{rm['bad_debt']}-{cl}{rm['nru_loss']}+{cl}{rm['other_income']}"), style='total')
    r += 2

    put(ws, r, 1, "Operating Expenses", style='section'); r += 1
    for key, lbl, curve in EXP:
        put(ws, r, 1, lbl, style='label')
        rm[key] = r
        for y in range(hold):
            cell = put(ws, r, 2 + y, _line(ws, r, y, key, curve, years, curves, rm, ti_row), style='money')
            ov = f"{key}:{y}"
            if ov in overrides and y < len(years) and years[y].computed_values:
                cell.value = overrides[ov]
                override(cell, years[y].computed_values.get(key, 0.0))
        r += 1

    te_row = r; rm["total_expenses"] = r
    put(ws, r, 1, "Total Operating Expenses", style='label_bold')
    for y in range(hold):
        cl = get_column_letter(2 + y)
        refs = "+".join(f"{cl}{rm[k]}" for k, *_ in EXP)
        put(ws, r, 2 + y, f"={refs}", style='total')
    r += 2

    rm["noi"] = r
    put(ws, r, 1, "Net Operating Income", style='label_bold')
    for y in range(hold):
        cl = get_column_letter(2 + y)
        put(ws, r, 2 + y, f"={cl}{ti_row}-{cl}{te_row}", style='total')
    r += 1

    rm["reserves"] = r
    put(ws, r, 1, "Capital Reserves", style='label')
    for y in range(hold):
        if y == 0:
            f = "=ReservesPerUnit*TotalUnits"
        elif inputs.reserves_inflate:
            f = f"={get_column_letter(2 + y - 1)}{r}*(1+{curves.get('ExpenseGrowth', [None] * (y + 1))[y]})"
        else:
            f = f"={get_column_letter(2)}{r}"
        put(ws, r, 2 + y, f, style='money')
    r += 1

    rm["ncf"] = r
    put(ws, r, 1, "Cash Flow from Operations", style='label_bold')
    for y in range(hold):
        cl = get_column_letter(2 + y)
        put(ws, r, 2 + y, f"={cl}{rm['noi']}-{cl}{rm['reserves']}", style='total')

    ws.freeze_panes = 'B4'
    return {"sheet": SHEET, "row_map": rm, "hold": hold}
