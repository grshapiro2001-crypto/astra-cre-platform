"""Cash Flows sheet — levered cash flow waterfall with IRR, EM, and CoC formulas."""

from openpyxl.utils import get_column_letter

from app.schemas.underwriting import UWInputs, ScenarioResult

from .styles import put

SHEET = "Cash Flows"


def build(wb, inputs: UWInputs, scenario: ScenarioResult, proforma_refs: dict, debt_refs: dict) -> dict:
    ws = wb.create_sheet(SHEET)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 32
    hold = proforma_refs["hold"]
    pf = proforma_refs["sheet"]
    rm = proforma_refs["row_map"]
    for c in range(2, 3 + hold):
        ws.column_dimensions[get_column_letter(c)].width = 15

    put(ws, 1, 1, "Levered Cash Flows", style='label_bold')
    put(ws, 3, 1, "", style='header')
    put(ws, 3, 2, "Year 0", style='header')
    for y in range(1, hold + 1):
        put(ws, 3, 2 + y, f"Year {y}", style='header')

    # Operating cash flow (NOI - DS - Reserves)
    put(ws, 4, 1, "NOI", style='label')
    put(ws, 5, 1, "Debt Service", style='label')
    put(ws, 6, 1, "Capital Reserves", style='label')
    put(ws, 7, 1, "Net Reversion Proceeds", style='label')
    for y in range(hold):
        pf_col = get_column_letter(2 + y)
        put(ws, 4, 3 + y, f"='{pf}'!{pf_col}{rm['noi']}", style='money')
        put(ws, 5, 3 + y, f"=-{debt_refs['ds_refs'][y]}", style='money')
        put(ws, 6, 3 + y, f"=-'{pf}'!{pf_col}{rm['reserves']}", style='money')
        put(ws, 7, 3 + y, 0, style='money')

    # Reversion in final year: GSP*(1-SaleCost) - Loan Balance, referencing engine GSP
    gsp = scenario.returns.reversion.gross_selling_price
    put(ws, 7, 3 + hold - 1,
        f"={gsp}*(1-SaleCostPct)-{debt_refs['end_bal_refs'][-1]}", style='money')

    # Annual operating cash-on-cash (NOI - DS - Reserves) / Equity
    put(ws, 9, 1, "Operating Cash Flow", style='label_bold')
    for y in range(hold):
        col = get_column_letter(3 + y)
        put(ws, 9, 3 + y, f"={col}4+{col}5+{col}6", style='total')

    # Total cash flow including acquisition + reversion
    put(ws, 10, 1, "Levered CF (incl. Reversion)", style='label_bold')
    put(ws, 10, 2, "=-Equity", style='total')
    for y in range(hold):
        col = get_column_letter(3 + y)
        put(ws, 10, 3 + y, f"={col}9+{col}7", style='total')

    first, last = get_column_letter(2), get_column_letter(2 + hold)
    op_first, op_last = get_column_letter(3), get_column_letter(2 + hold)

    put(ws, 12, 1, "Levered IRR", style='label_bold')
    put(ws, 12, 2, f"=IRR({first}10:{last}10)", style='percent_precise')

    put(ws, 13, 1, "Equity Multiple", style='label_bold')
    put(ws, 13, 2,
        f"=(SUMIF({op_first}10:{last}10,\">0\")+Equity)/Equity", style='multiple')

    put(ws, 14, 1, "Avg Cash-on-Cash", style='label_bold')
    put(ws, 14, 2, f"=AVERAGE({op_first}9:{op_last}9)/Equity", style='percent_precise')

    ws.freeze_panes = 'C4'
    return {
        "sheet": SHEET,
        "irr_cell": f"'{SHEET}'!$B$12",
        "em_cell": f"'{SHEET}'!$B$13",
        "coc_cell": f"'{SHEET}'!$B$14",
    }
