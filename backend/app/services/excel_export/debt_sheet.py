"""Debt sheet — annual amortization schedule keyed off Assumptions named ranges."""

from openpyxl.utils import get_column_letter

from app.schemas.underwriting import UWInputs, ScenarioResult

from .styles import put

SHEET = "Debt"


def build(wb, inputs: UWInputs, scenario: ScenarioResult) -> dict:
    ws = wb.create_sheet(SHEET)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 8
    for c in range(2, 8):
        ws.column_dimensions[get_column_letter(c)].width = 16

    put(ws, 1, 1, "Debt — Annual Amortization Schedule", style='label_bold')
    hold = max(1, inputs.hold_period_years)

    headers = ["Year", "Beg Balance", "Interest", "Principal", "End Balance", "Annual DS", "DSCR"]
    for i, h in enumerate(headers):
        put(ws, 3, 1 + i, h, style='header')

    annual_ds = scenario.debt.annual_debt_service or [0.0] * hold
    end_balances = scenario.debt.principal_outstanding or [scenario.debt.loan_amount] * hold

    ds_refs: list[str] = []
    end_bal_refs: list[str] = []

    for y in range(1, hold + 1):
        r = 3 + y
        put(ws, r, 1, y, style='label')

        beg_formula = "=LoanAmount" if y == 1 else f"=E{r - 1}"
        put(ws, r, 2, beg_formula, style='money')
        put(ws, r, 3, f"=B{r}*IntRate", style='money')
        put(ws, r, 4, f"=F{r}-C{r}", style='money')
        put(ws, r, 5, end_balances[y - 1] if y - 1 < len(end_balances) else 0.0, style='money')
        put(ws, r, 6, annual_ds[y - 1] if y - 1 < len(annual_ds) else 0.0, style='money')
        put(ws, r, 7, "", style='money')

        ds_refs.append(f"'{SHEET}'!$F${r}")
        end_bal_refs.append(f"'{SHEET}'!$E${r}")

    ws.freeze_panes = 'B4'

    return {
        "sheet": SHEET,
        "ds_refs": ds_refs,
        "end_bal_refs": end_bal_refs,
        "ds_first_row": 4,
        "ds_last_row": 3 + hold,
    }
