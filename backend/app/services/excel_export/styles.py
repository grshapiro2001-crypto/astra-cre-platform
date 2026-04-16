"""Cell styles, number formats, colors, and fonts for the UW export."""

from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

MONEY_FMT = '#,##0;[Red](#,##0);"—"'
PCT_FMT = '0.0%;[Red](0.0%)'
PCT_PRECISE_FMT = '0.00%;[Red](0.00%)'
MULTIPLE_FMT = '0.00"x"'

HEADER_FONT = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
HEADER_FILL = PatternFill('solid', fgColor='1F2A44')
SECTION_FONT = Font(name='Calibri', size=11, bold=True, color='1F2A44')
SECTION_FILL = PatternFill('solid', fgColor='EAEEF7')
TITLE_FONT = Font(name='Calibri', size=14, bold=True, color='1F2A44')
LABEL_FONT = Font(name='Calibri', size=10, bold=True)
BODY_FONT = Font(name='Calibri', size=10)
TOTAL_FONT = Font(name='Calibri', size=10, bold=True)

OVERRIDE_FILL = PatternFill('solid', fgColor='FFF3BF')
INPUT_FILL = PatternFill('solid', fgColor='FBFAED')

TOTAL_BORDER = Border(top=Side(style='thin'), bottom=Side(style='double'))
CENTER = Alignment(horizontal='center', vertical='center')
RIGHT = Alignment(horizontal='right', vertical='center')
LEFT = Alignment(horizontal='left', vertical='center')


def override(cell, computed_value: float) -> None:
    """Mark a cell as an override: yellow fill + comment noting model value."""
    cell.fill = OVERRIDE_FILL
    cell.comment = Comment(
        f"Override. Model computed: {computed_value:,.2f}", "Astra Underwriting",
    )


_STYLES = {
    'header':         (HEADER_FONT, HEADER_FILL, CENTER, None, None),
    'section':        (SECTION_FONT, SECTION_FILL, LEFT, None, None),
    'title':          (TITLE_FONT, None, LEFT, None, None),
    'label':          (BODY_FONT, None, LEFT, None, None),
    'label_bold':     (LABEL_FONT, None, LEFT, None, None),
    'money':          (BODY_FONT, None, RIGHT, MONEY_FMT, None),
    'percent':        (BODY_FONT, None, RIGHT, PCT_FMT, None),
    'percent_precise': (BODY_FONT, None, RIGHT, PCT_PRECISE_FMT, None),
    'multiple':       (BODY_FONT, None, RIGHT, MULTIPLE_FMT, None),
    'total':          (TOTAL_FONT, None, RIGHT, MONEY_FMT, TOTAL_BORDER),
    'input':          (BODY_FONT, INPUT_FILL, LEFT, None, None),
    'input_money':    (BODY_FONT, INPUT_FILL, RIGHT, MONEY_FMT, None),
    'input_pct':      (BODY_FONT, INPUT_FILL, RIGHT, PCT_PRECISE_FMT, None),
}


def put(ws, row: int, col: int, value, *, style: str | None = None):
    cell = ws.cell(row=row, column=col, value=value)
    spec = _STYLES.get(style)
    if spec:
        font, fill, align, fmt, border = spec
        if font:   cell.font = font
        if fill:   cell.fill = fill
        if align:  cell.alignment = align
        if fmt:    cell.number_format = fmt
        if border: cell.border = border
    return cell
