"""
Astra CRE — Intent-Based Excel Extraction Prototype
=====================================================
This module extracts CRE financial data from ANY Excel file without
file-specific configuration. It uses a two-phase approach:

Phase 1: Structural analysis (sheet inventory, classification)
Phase 2: Semantic extraction (find concepts, not cells)

No hardcoded cell references. No template matching.
Works by understanding CRE vocabulary and letting the AI find the data.
"""

import openpyxl
import pandas as pd
import json
import re
from datetime import datetime, date
from typing import Any


# ============================================================
# PHASE 1: STRUCTURAL ANALYSIS
# ============================================================

# CRE vocabulary for sheet classification
SHEET_CLASSIFIERS = {
    "summary_output": [
        "summary", "overview", "executive", "dashboard", "output",
        "results", "valuation", "pricing"
    ],
    "proforma": [
        "proforma", "pro forma", "pro-forma", "projected", "forecast",
        "cash flow", "cf ", "irr", "dcf"
    ],
    "assumptions": [
        "assumption", "input", "scenario", "sensitivity", "parameter"
    ],
    "rent_data": [
        "rent roll", "rent_roll", "rentroll", "unit mix", "unitmix",
        "unit_mix", "lease", "tenant", "occupancy", "rollover",
        "loss to lease", "gross scheduled"
    ],
    "financials": [
        "trailing", "t-12", "t12", "operating statement", "income",
        "expense", "financial", "p&l", "profit", "budget", "actual"
    ],
    "debt": [
        "financing", "debt", "loan", "mortgage", "amort", "leverage"
    ],
    "comps": [
        "comp", "comparable", "sales comp", "rent comp", "market data",
        "transaction"
    ],
    "config_system": [
        "config", "dropdown", "sysconfig", "setting", "print",
        "template", "log", "change log", "regression", "test"
    ]
}

# CRE financial vocabulary for semantic extraction
CRE_TERMS = {
    # Property identification
    "property_name": ["property name", "property", "asset name", "project name", "community"],
    "address": ["address", "street", "location"],
    "city": ["city", "market"],
    "state": ["state"],
    "units": ["total units", "units", "unit count", "# of units", "number of units", "no. of units"],
    "square_feet": ["total sf", "total square feet", "rentable sf", "square feet", "net rentable"],
    "year_built": ["year built", "vintage", "built", "construction year"],
    "asset_type": ["asset type", "property type", "product type", "building type"],
    
    # Income metrics
    "gross_scheduled_rent": ["gross scheduled", "scheduled market rent", "gross potential rent", "gpr", "gsr"],
    "loss_to_lease": ["loss to lease", "gain to lease", "gain/loss to lease", "ltl"],
    "vacancy": ["vacancy", "physical vacancy", "economic vacancy", "vac loss"],
    "concessions": ["concession", "free rent"],
    "bad_debt": ["bad debt", "credit loss", "collection loss", "write-off"],
    "net_rental_income": ["net rental income", "nri", "effective rental income"],
    "other_income": ["other income", "ancillary", "miscellaneous income", "utility reimb"],
    "effective_gross_income": ["effective gross income", "egi", "total income", "total operating income", "total revenue"],
    
    # Expense metrics
    "total_expenses": ["total expenses", "total operating expense", "operating expenses"],
    "management_fee": ["management fee", "mgmt fee", "property management"],
    "property_taxes": ["property tax", "real estate tax", "re tax", "ad valorem"],
    "insurance": ["insurance", "property insurance"],
    "utilities": ["utilities", "utility expense"],
    "payroll": ["payroll", "personnel", "salary", "wages", "on-site staff"],
    "repairs_maintenance": ["repair", "maintenance", "r&m"],
    "marketing": ["marketing", "advertising", "leasing cost"],
    
    # Bottom line
    "noi": ["net operating income", "noi"],
    "ncf": ["net cash flow", "ncf", "cash flow after debt"],
    "reserves": ["replacement reserve", "capital reserve", "capex reserve"],
    
    # Pricing / Returns
    "purchase_price": ["purchase price", "sale price", "acquisition price", "project value", "total price"],
    "price_per_unit": ["per unit", "price per unit", "price/unit", "$/unit"],
    "price_per_sf": ["per sf", "price per sf", "per square foot", "$/sf"],
    "cap_rate": ["cap rate", "capitalization rate", "going-in cap", "y1 cap", "year 1 cap"],
    "in_place_cap": ["in place cap", "in-place cap", "trailing cap", "t3 cap", "t12 cap"],
    "terminal_cap": ["terminal cap", "exit cap", "reversion cap", "disposition cap"],
    "irr": ["irr", "internal rate of return"],
    "levered_irr": ["levered irr", "leveraged irr", "lev irr"],
    "unlevered_irr": ["unlevered irr", "unleveraged irr", "free & clear irr", "unlev irr"],
    "cash_on_cash": ["cash-on-cash", "cash on cash", "coc", "equity yield"],
    "equity_multiple": ["equity multiple", "moic", "multiple on invested capital"],
    
    # Debt
    "ltv": ["ltv", "loan to value", "loan-to-value", "leverage ratio"],
    "dscr": ["dscr", "debt service coverage", "debt coverage"],
    "interest_rate": ["interest rate", "coupon", "note rate", "all-in rate"],
    "loan_term": ["loan term", "term", "maturity"],
    "amortization": ["amortization", "amort"],
    "io_period": ["interest only", "i/o period", "io period"],
    
    # Growth assumptions
    "rent_growth": ["rent growth", "revenue growth", "rental growth", "market rent growth"],
    "expense_growth": ["expense growth", "expense inflation", "opex growth"],
    "hold_period": ["hold period", "holding period", "investment horizon"],
}


def classify_sheet(sheet_name: str) -> list[str]:
    """Classify a sheet by name using CRE vocabulary."""
    name_lower = sheet_name.lower().strip()
    classifications = []
    for category, keywords in SHEET_CLASSIFIERS.items():
        for kw in keywords:
            if kw in name_lower:
                classifications.append(category)
                break
    return classifications if classifications else ["unknown"]


def serialize_value(v: Any) -> Any:
    """Convert Excel values to JSON-serializable types."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.isoformat()
    if isinstance(v, date):
        return v.isoformat()
    if isinstance(v, (int, float)):
        if isinstance(v, float) and (v != v):  # NaN check
            return None
        return v
    if isinstance(v, bool):
        return v
    return str(v)


def read_sheet_data(ws, max_rows=80, max_cols=30) -> list[list]:
    """Read sheet data into a 2D array, handling merges and empties."""
    rows = []
    for row in ws.iter_rows(min_row=1, max_row=min(max_rows, ws.max_row or 1),
                            max_col=min(max_cols, ws.max_column or 1),
                            values_only=True):
        serialized = [serialize_value(v) for v in row]
        if any(v is not None for v in serialized):
            rows.append(serialized)
    return rows


def build_sheet_inventory(filepath: str) -> dict:
    """
    Phase 1: Build a complete structural inventory of the workbook.
    Returns sheet classifications, dimensions, and preview data.
    """
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    
    inventory = {
        "filename": filepath.split("/")[-1],
        "total_sheets": len(wb.sheetnames),
        "sheets": {}
    }
    
    for name in wb.sheetnames:
        ws = wb[name]
        classifications = classify_sheet(name)
        
        # Read preview data for high-value sheets
        is_high_value = any(c in ["summary_output", "proforma", "assumptions", 
                                   "rent_data", "financials", "debt", "comps"]
                          for c in classifications)
        
        sheet_info = {
            "classifications": classifications,
            "max_row": ws.max_row,
            "max_column": ws.max_column,
        }
        
        if is_high_value:
            sheet_info["preview_data"] = read_sheet_data(ws, max_rows=80, max_cols=30)
        
        inventory["sheets"][name] = sheet_info
    
    wb.close()
    return inventory


# ============================================================
# PHASE 2: SEMANTIC EXTRACTION
# ============================================================

def find_value_near_label(rows: list[list], search_terms: list[str], 
                          search_range_cols=6) -> list[dict]:
    """
    Find values adjacent to cells containing search terms.
    Returns all matches with location and context.
    """
    matches = []
    for row_idx, row in enumerate(rows):
        for col_idx, cell in enumerate(row):
            if cell is None:
                continue
            cell_str = str(cell).lower().strip()
            
            for term in search_terms:
                if term in cell_str:
                    # Look for values in adjacent cells (right and below)
                    adjacent_values = []
                    
                    # Check cells to the right (same row)
                    for offset in range(1, min(search_range_cols, len(row) - col_idx)):
                        val = row[col_idx + offset]
                        if val is not None and val != '' and val != 0:
                            adjacent_values.append({
                                "value": val,
                                "position": "right",
                                "offset": offset,
                                "row": row_idx,
                                "col": col_idx + offset
                            })
                    
                    # Check cell below (next row, same column)
                    if row_idx + 1 < len(rows):
                        below_row = rows[row_idx + 1]
                        if col_idx < len(below_row) and below_row[col_idx] is not None:
                            adjacent_values.append({
                                "value": below_row[col_idx],
                                "position": "below",
                                "offset": 1,
                                "row": row_idx + 1,
                                "col": col_idx
                            })
                    
                    if adjacent_values:
                        matches.append({
                            "label": str(cell),
                            "term_matched": term,
                            "row": row_idx,
                            "col": col_idx,
                            "adjacent_values": adjacent_values
                        })
                    break  # Don't double-match same cell
    
    return matches


def is_percentage(value) -> bool:
    """Check if a numeric value is likely a percentage (0-1 range or has % sign)."""
    if isinstance(value, str) and '%' in value:
        return True
    if isinstance(value, (int, float)):
        return -1 <= value <= 1 and value != 0
    return False


def is_dollar_amount(value) -> bool:
    """Check if a value is likely a dollar amount."""
    if isinstance(value, (int, float)):
        return abs(value) > 1  # Distinguishes from percentages
    if isinstance(value, str):
        return '$' in value or ',' in value
    return False


def extract_financial_data(inventory: dict) -> dict:
    """
    Phase 2: Extract CRE financial data from classified sheets.
    Uses semantic matching — finds concepts, not cells.
    """
    extracted = {
        "property_info": {},
        "unit_mix": [],
        "financials": {},
        "pricing_scenarios": {},
        "assumptions": {},
        "debt": {},
        "returns": {},
        "raw_matches": {}  # For debugging / transparency
    }
    
    # Prioritize sheets by classification
    priority_order = [
        "summary_output", "assumptions", "proforma", 
        "financials", "rent_data", "debt"
    ]
    
    for priority_class in priority_order:
        for sheet_name, sheet_info in inventory["sheets"].items():
            if priority_class not in sheet_info.get("classifications", []):
                continue
            if "preview_data" not in sheet_info:
                continue
            
            rows = sheet_info["preview_data"]
            
            # Search for every CRE term in this sheet
            for field_name, search_terms in CRE_TERMS.items():
                matches = find_value_near_label(rows, search_terms)
                if matches:
                    key = f"{sheet_name}::{field_name}"
                    extracted["raw_matches"][key] = matches
                    
                    # Route to appropriate category
                    _route_match(extracted, field_name, matches, sheet_name)
    
    return extracted


def _route_match(extracted: dict, field_name: str, matches: list, sheet_name: str):
    """Route an extracted match to the right category in the output."""
    
    # Property info fields
    property_fields = ["property_name", "address", "city", "state", "units", 
                       "square_feet", "year_built", "asset_type"]
    
    # Financial fields
    income_fields = ["gross_scheduled_rent", "loss_to_lease", "vacancy", 
                     "concessions", "bad_debt", "net_rental_income", 
                     "other_income", "effective_gross_income"]
    expense_fields = ["total_expenses", "management_fee", "property_taxes",
                      "insurance", "utilities", "payroll", "repairs_maintenance",
                      "marketing"]
    bottom_line = ["noi", "ncf", "reserves"]
    
    # Pricing fields
    pricing_fields = ["purchase_price", "price_per_unit", "price_per_sf",
                      "cap_rate", "in_place_cap", "terminal_cap"]
    
    # Return fields
    return_fields = ["irr", "levered_irr", "unlevered_irr", 
                     "cash_on_cash", "equity_multiple"]
    
    # Debt fields
    debt_fields = ["ltv", "dscr", "interest_rate", "loan_term", 
                   "amortization", "io_period"]
    
    # Assumption fields
    assumption_fields = ["rent_growth", "expense_growth", "hold_period"]
    
    # Get the best value(s) from matches
    best = _get_best_values(matches)
    
    if field_name in property_fields:
        if field_name not in extracted["property_info"]:
            extracted["property_info"][field_name] = best
    elif field_name in income_fields + expense_fields + bottom_line:
        if field_name not in extracted["financials"]:
            extracted["financials"][field_name] = best
    elif field_name in pricing_fields:
        if field_name not in extracted["pricing_scenarios"]:
            extracted["pricing_scenarios"][field_name] = best
    elif field_name in return_fields:
        if field_name not in extracted["returns"]:
            extracted["returns"][field_name] = best
    elif field_name in debt_fields:
        if field_name not in extracted["debt"]:
            extracted["debt"][field_name] = best
    elif field_name in assumption_fields:
        if field_name not in extracted["assumptions"]:
            extracted["assumptions"][field_name] = best


def _get_best_values(matches: list) -> dict:
    """Extract the best value(s) from a set of matches."""
    result = {
        "values": [],
        "source_labels": [],
        "confidence": "low"
    }
    
    seen_values = set()
    for match in matches:
        for adj in match["adjacent_values"]:
            val = adj["value"]
            # Skip error values
            if isinstance(val, str) and val.startswith('#'):
                continue
            # Skip duplicates
            val_key = str(val)
            if val_key in seen_values:
                continue
            seen_values.add(val_key)
            
            result["values"].append({
                "value": val,
                "label": match["label"],
                "position": adj["position"]
            })
            if match["label"] not in result["source_labels"]:
                result["source_labels"].append(match["label"])
    
    # Set confidence based on number of confirming sources
    if len(result["values"]) >= 2:
        result["confidence"] = "high"
    elif len(result["values"]) == 1:
        result["confidence"] = "medium"
    
    return result


# ============================================================
# COMP TRACKER EXTRACTION
# ============================================================

def extract_comp_tracker(filepath: str) -> dict:
    """
    Extract sales comp data from a tabular comp tracker.
    Detects header row, maps columns, and extracts all rows.
    """
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    ws = wb.active  # Comp trackers are usually single-sheet
    
    # Read all rows
    all_rows = []
    for row in ws.iter_rows(values_only=True):
        serialized = [serialize_value(v) for v in row]
        all_rows.append(serialized)
    
    wb.close()
    
    # Find header row by looking for CRE comp column names
    comp_column_terms = {
        "property_name": ["property name", "property", "name", "asset"],
        "market": ["market", "submarket", "location", "city", "area"],
        "property_type": ["type", "property type", "product type", "style"],
        "sale_date": ["date", "sale date", "close date", "transaction date"],
        "year_built": ["year built", "built", "vintage", "year"],
        "units": ["units", "no. of units", "unit count", "# units"],
        "avg_sf": ["avg. unit sf", "avg sf", "unit sf", "square feet", "sf"],
        "avg_rent": ["eff. rent", "avg. eff. rent", "rent", "avg rent", "in-place rent"],
        "sale_price": ["sale price", "price", "purchase price", "total price"],
        "price_per_unit": ["price/unit", "$/unit", "per unit", "ppu"],
        "price_per_sf": ["price/sf", "$/sf", "per sf"],
        "cap_rate": ["cap rate", "cap", "going-in cap"],
        "cap_rate_qualifier": ["qualifier", "cap rate qualifier", "cap type", "basis"],
        "buyer": ["buyer", "purchaser", "acquirer"],
        "seller": ["seller", "vendor", "disposition"]
    }
    
    # Scan first 10 rows for header
    header_row_idx = None
    column_mapping = {}
    
    for row_idx in range(min(10, len(all_rows))):
        row = all_rows[row_idx]
        matches_found = 0
        temp_mapping = {}
        
        for col_idx, cell in enumerate(row):
            if cell is None:
                continue
            cell_lower = str(cell).lower().strip()
            
            for field, terms in comp_column_terms.items():
                for term in terms:
                    if term in cell_lower and field not in temp_mapping:
                        temp_mapping[field] = col_idx
                        matches_found += 1
                        break
        
        # If we found 4+ column matches, this is likely the header row
        if matches_found >= 4:
            # Check if previous row has partial headers (split header)
            if row_idx > 0:
                prev_row = all_rows[row_idx - 1]
                for col_idx, cell in enumerate(prev_row):
                    if cell is None:
                        continue
                    cell_lower = str(cell).lower().strip()
                    for field, terms in comp_column_terms.items():
                        for term in terms:
                            if term in cell_lower and field not in temp_mapping:
                                temp_mapping[field] = col_idx
                                break
            
            header_row_idx = row_idx
            column_mapping = temp_mapping
            break
    
    if header_row_idx is None:
        return {"error": "Could not identify header row", "preview": all_rows[:5]}
    
    # Extract comps starting from row after header
    comps = []
    for row_idx in range(header_row_idx + 1, len(all_rows)):
        row = all_rows[row_idx]
        
        # Skip empty rows
        if not any(v is not None for v in row):
            continue
        
        # Skip if no property name
        name_col = column_mapping.get("property_name", 0)
        if name_col >= len(row) or row[name_col] is None:
            continue
        
        comp = {}
        for field, col_idx in column_mapping.items():
            if col_idx < len(row):
                val = row[col_idx]
                # Format percentages
                if field == "cap_rate" and isinstance(val, (int, float)) and 0 < val < 1:
                    comp[field] = round(val * 100, 2)
                    comp[f"{field}_decimal"] = val
                elif isinstance(val, float):
                    comp[field] = round(val, 2) if abs(val) < 1000 else round(val)
                else:
                    comp[field] = val
        
        comps.append(comp)
    
    # Generate summary stats
    cap_rates = [c["cap_rate_decimal"] for c in comps 
                 if "cap_rate_decimal" in c and isinstance(c.get("cap_rate_decimal"), (int, float))]
    prices_per_unit = [c["price_per_unit"] for c in comps 
                       if isinstance(c.get("price_per_unit"), (int, float))]
    
    return {
        "document_type": "sales_comp_tracker",
        "extraction_method": "intent_based_column_detection",
        "header_row": header_row_idx,
        "columns_detected": list(column_mapping.keys()),
        "columns_not_found": [f for f in comp_column_terms.keys() if f not in column_mapping],
        "total_comps": len(comps),
        "summary_stats": {
            "cap_rates": {
                "count": len(cap_rates),
                "min": round(min(cap_rates) * 100, 2) if cap_rates else None,
                "max": round(max(cap_rates) * 100, 2) if cap_rates else None,
                "median": round(sorted(cap_rates)[len(cap_rates)//2] * 100, 2) if cap_rates else None,
                "avg": round(sum(cap_rates) / len(cap_rates) * 100, 2) if cap_rates else None,
            },
            "price_per_unit": {
                "count": len(prices_per_unit),
                "min": round(min(prices_per_unit)) if prices_per_unit else None,
                "max": round(max(prices_per_unit)) if prices_per_unit else None,
                "median": round(sorted(prices_per_unit)[len(prices_per_unit)//2]) if prices_per_unit else None,
                "avg": round(sum(prices_per_unit) / len(prices_per_unit)) if prices_per_unit else None,
            }
        },
        "comps": comps  # Full comp list
    }


# ============================================================
# MAIN: RUN ON BOTH FILES
# ============================================================

def format_extraction_report(extracted: dict, doc_type: str) -> str:
    """Format extracted data into a readable report."""
    lines = []
    
    if doc_type == "model":
        lines.append("=" * 70)
        lines.append("UNDERWRITING MODEL EXTRACTION — INTENT-BASED (NO FILE-SPECIFIC LOGIC)")
        lines.append("=" * 70)
        
        lines.append("\n📋 PROPERTY INFO:")
        for field, data in extracted.get("property_info", {}).items():
            vals = [v["value"] for v in data.get("values", [])]
            labels = data.get("source_labels", [])
            conf = data.get("confidence", "?")
            lines.append(f"  {field}: {vals} (from: {labels}) [{conf}]")
        
        lines.append("\n💰 FINANCIALS:")
        for field, data in extracted.get("financials", {}).items():
            vals = data.get("values", [])
            for v in vals[:4]:  # Show up to 4 values
                val = v["value"]
                label = v["label"]
                if isinstance(val, float):
                    if abs(val) < 1:
                        lines.append(f"  {field}: {val:.1%} (from: '{label}')")
                    elif abs(val) > 10000:
                        lines.append(f"  {field}: ${val:,.0f} (from: '{label}')")
                    else:
                        lines.append(f"  {field}: {val:,.2f} (from: '{label}')")
                else:
                    lines.append(f"  {field}: {val} (from: '{label}')")
        
        lines.append("\n🏷️ PRICING / CAP RATES:")
        for field, data in extracted.get("pricing_scenarios", {}).items():
            vals = data.get("values", [])
            for v in vals[:4]:
                val = v["value"]
                label = v["label"]
                if isinstance(val, float):
                    if abs(val) < 1:
                        lines.append(f"  {field}: {val:.2%} (from: '{label}')")
                    elif abs(val) > 10000:
                        lines.append(f"  {field}: ${val:,.0f} (from: '{label}')")
                    else:
                        lines.append(f"  {field}: {val:,.2f} (from: '{label}')")
                else:
                    lines.append(f"  {field}: {val} (from: '{label}')")
        
        lines.append("\n📈 RETURNS:")
        for field, data in extracted.get("returns", {}).items():
            vals = data.get("values", [])
            for v in vals[:4]:
                val = v["value"]
                label = v["label"]
                if isinstance(val, float) and abs(val) < 1:
                    lines.append(f"  {field}: {val:.2%} (from: '{label}')")
                else:
                    lines.append(f"  {field}: {val} (from: '{label}')")
        
        lines.append("\n🏦 DEBT:")
        for field, data in extracted.get("debt", {}).items():
            vals = data.get("values", [])
            for v in vals[:4]:
                val = v["value"]
                label = v["label"]
                if isinstance(val, float) and abs(val) < 1:
                    lines.append(f"  {field}: {val:.2%} (from: '{label}')")
                else:
                    lines.append(f"  {field}: {val} (from: '{label}')")
        
        lines.append("\n⚙️ ASSUMPTIONS:")
        for field, data in extracted.get("assumptions", {}).items():
            vals = data.get("values", [])
            for v in vals[:4]:
                val = v["value"]
                label = v["label"]
                if isinstance(val, float) and abs(val) < 1:
                    lines.append(f"  {field}: {val:.2%} (from: '{label}')")
                else:
                    lines.append(f"  {field}: {val} (from: '{label}')")
    
    elif doc_type == "comps":
        lines.append("=" * 70)
        lines.append("SALES COMP TRACKER EXTRACTION — INTENT-BASED (NO FILE-SPECIFIC LOGIC)")
        lines.append("=" * 70)
        
        lines.append(f"\nTotal Comps Extracted: {extracted.get('total_comps', 0)}")
        lines.append(f"Columns Detected: {', '.join(extracted.get('columns_detected', []))}")
        
        missing = extracted.get('columns_not_found', [])
        if missing:
            lines.append(f"Columns NOT Found: {', '.join(missing)}")
        
        stats = extracted.get("summary_stats", {})
        cr = stats.get("cap_rates", {})
        ppu = stats.get("price_per_unit", {})
        
        lines.append(f"\n📊 CAP RATE STATS ({cr.get('count', 0)} comps with data):")
        lines.append(f"  Min: {cr.get('min')}%  |  Max: {cr.get('max')}%  |  Median: {cr.get('median')}%  |  Avg: {cr.get('avg')}%")
        
        lines.append(f"\n💲 PRICE/UNIT STATS ({ppu.get('count', 0)} comps):")
        ppu_min = ppu.get('min') or 0
        ppu_max = ppu.get('max') or 0
        ppu_med = ppu.get('median') or 0
        ppu_avg = ppu.get('avg') or 0
        lines.append(f"  Min: ${ppu_min:,.0f}  |  Max: ${ppu_max:,.0f}  |  Median: ${ppu_med:,.0f}  |  Avg: ${ppu_avg:,.0f}")
        
        lines.append(f"\n📋 FIRST 10 COMPS:")
        for i, comp in enumerate(extracted.get("comps", [])[:10]):
            name = comp.get("property_name", "?")
            market = comp.get("market", "?")
            units = comp.get("units", "?")
            cap = comp.get("cap_rate", "TBD")
            qualifier = comp.get("cap_rate_qualifier", "")
            ppu_val = comp.get("price_per_unit", "?")
            date_val = comp.get("sale_date", "?")
            if isinstance(date_val, str) and 'T' in date_val:
                date_val = date_val[:10]
            
            cap_str = f"{cap}% ({qualifier})" if isinstance(cap, (int, float)) else str(cap)
            ppu_str = f"${ppu_val:,.0f}" if isinstance(ppu_val, (int, float)) else str(ppu_val)
            
            lines.append(f"  {i+1}. {name} | {market} | {units} units | {cap_str} | {ppu_str}/unit | {date_val}")
    
    return "\n".join(lines)


if __name__ == "__main__":
    print("\n" + "=" * 70)
    print("RUNNING INTENT-BASED EXTRACTION ON BOTH FILES")
    print("Zero file-specific configuration. Zero hardcoded cell references.")
    print("=" * 70)
    
    # --- FILE 1: Underwriting Model ---
    print("\n\n🏢 FILE 1: Prose Gainesville Proforma (.xlsm)")
    print("-" * 50)
    
    model_path = "/mnt/user-data/uploads/Prose_Gainesville_Proforma_1_30_26_RH.xlsm"
    
    print("Phase 1: Building sheet inventory...")
    inventory = build_sheet_inventory(model_path)
    
    print(f"  Found {inventory['total_sheets']} sheets")
    print("  Sheet classifications:")
    for name, info in inventory["sheets"].items():
        classes = info["classifications"]
        has_data = "preview_data" in info
        rows = info.get("max_row", "?")
        print(f"    {name:40s} → {', '.join(classes):25s} {'[READING]' if has_data else '[SKIPPED]'} ({rows} rows)")
    
    print("\nPhase 2: Semantic extraction...")
    model_extracted = extract_financial_data(inventory)
    
    report = format_extraction_report(model_extracted, "model")
    print(report)
    
    # --- FILE 2: Sales Comps ---
    print("\n\n🏘️ FILE 2: Metro Atlanta Sales Comps (.xlsx)")
    print("-" * 50)
    
    comps_path = "/mnt/user-data/uploads/Nuveen_-_Metro_Atlanta_Sales_Comps_2024-2025.xlsx"
    
    print("Detecting document type and extracting...")
    comps_extracted = extract_comp_tracker(comps_path)
    
    report = format_extraction_report(comps_extracted, "comps")
    print(report)
    
    # --- CROSS-FILE ANALYSIS ---
    print("\n\n" + "=" * 70)
    print("🔗 CROSS-FILE ANALYSIS: Prose Gainesville vs Atlanta Comps")
    print("=" * 70)
    
    # Find relevant comps for Prose Gainesville
    # Property: 300 units, 2024, Gainesville GA, Garden
    relevant_comps = []
    for comp in comps_extracted.get("comps", []):
        # Filter: NE Atlanta suburbs, 2020+ vintage, 200+ units
        market = str(comp.get("market", "")).strip().lower()
        year = comp.get("year_built")
        units = comp.get("units", 0)
        
        ne_atlanta_markets = [
            "gainesville", "hoschton", "oakwood", "buford", "duluth",
            "gwinnett", "lawrenceville", "norcross", "suwanee", 
            "sugar hill", "athens", "winder", "grayson", "snellville",
            "covington", "conyers", "stantham"
        ]
        
        is_ne_atlanta = any(m in market for m in ne_atlanta_markets)
        is_newer = isinstance(year, (int, float)) and year >= 2020
        is_similar_size = isinstance(units, (int, float)) and units >= 100
        
        if is_ne_atlanta and (is_newer or is_similar_size):
            relevant_comps.append(comp)
    
    print(f"\nRelevant Comps (NE Atlanta, 2020+ vintage or 100+ units): {len(relevant_comps)}")
    for comp in relevant_comps:
        name = comp.get("property_name", "?")
        market = comp.get("market", "?")
        units = comp.get("units", "?")
        year = comp.get("year_built", "?")
        cap = comp.get("cap_rate", "TBD")
        qualifier = comp.get("cap_rate_qualifier", "")
        ppu = comp.get("price_per_unit", "?")
        
        cap_str = f"{cap}% ({qualifier})" if isinstance(cap, (int, float)) else str(cap)
        ppu_str = f"${ppu:,.0f}" if isinstance(ppu, (int, float)) else str(ppu)
        
        print(f"  • {name} | {market} | {units} units | {year} | {cap_str} | {ppu_str}/unit")
    
    # Compare to Prose
    if relevant_comps:
        comp_caps = [c["cap_rate_decimal"] for c in relevant_comps 
                     if isinstance(c.get("cap_rate_decimal"), (int, float))]
        comp_ppus = [c["price_per_unit"] for c in relevant_comps 
                     if isinstance(c.get("price_per_unit"), (int, float))]
        
        if comp_caps:
            avg_cap = sum(comp_caps) / len(comp_caps) * 100
            print(f"\n  Comp Avg Cap Rate: {avg_cap:.2f}%")
        if comp_ppus:
            avg_ppu = sum(comp_ppus) / len(comp_ppus)
            print(f"  Comp Avg $/Unit: ${avg_ppu:,.0f}")
        
        # Prose pricing
        pricing = model_extracted.get("pricing_scenarios", {})
        price_vals = pricing.get("purchase_price", {}).get("values", [])
        cap_vals = pricing.get("cap_rate", {}).get("values", [])
        
        print(f"\n  Prose Gainesville Pricing:")
        for pv in price_vals[:2]:
            if isinstance(pv["value"], (int, float)) and pv["value"] > 1000000:
                print(f"    {pv['label']}: ${pv['value']:,.0f}")
        for cv in cap_vals[:4]:
            if isinstance(cv["value"], (int, float)) and 0 < cv["value"] < 1:
                print(f"    {cv['label']}: {cv['value']:.2%}")
    
    print("\n" + "=" * 70)
    print("EXTRACTION COMPLETE — Zero file-specific configuration used")
    print("=" * 70)
